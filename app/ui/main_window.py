"""
Fenêtre principale de l'application
"""

import os
import sys
import tempfile
from pathlib import Path
from typing import Optional
from datetime import datetime, date, timedelta
from loguru import logger
from dotenv import load_dotenv
import pandas as pd

from PySide6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
    QTableWidget, QTableWidgetItem, QLabel, QComboBox, QDateEdit,
    QMessageBox, QDialog, QFileDialog, QStatusBar, QToolBar, QMenu,
    QScrollArea, QGridLayout, QFrame, QProgressBar, QSizePolicy, QCheckBox,
    QStyledItemDelegate, QStyle
)
from PySide6.QtCore import Qt, QTimer, QDate, Signal, Slot, QSize, QByteArray
from PySide6.QtGui import QAction, QIcon, QPixmap, QPainter, QBrush, QColor, QPen
from PySide6.QtSvg import QSvgRenderer

# Import des utilitaires
sys.path.append(str(Path(__file__).parent.parent.parent))
from app.utils import get_resource_path


class NoSelectionDelegate(QStyledItemDelegate):
    """Delegate qui utilise le background de l'item au lieu du style de sélection Qt"""
    def paint(self, painter, option, index):
        # Retirer le flag Selected pour empêcher Qt de dessiner son propre style
        option.state &= ~QStyle.State_Selected

        # Récupérer le background personnalisé de l'item (défini par setBackground)
        bg_data = index.data(Qt.BackgroundRole)
        if bg_data:
            option.backgroundBrush = QBrush(bg_data)

        # Dessiner l'item normalement avec notre background personnalisé
        super().paint(painter, option, index)

sys.path.append(str(Path(__file__).parent.parent.parent))

from app.services.supabase_client import supabase_client
from app.services.file_processor import FileProcessor
from app.models.file_record import FileRecord, FileStatus, FileType
from app.ui.transformation_config_dialog import TransformationConfigDialog
from worker.ftp_fetcher import FTPFetcher

load_dotenv()


class MainWindow(QMainWindow):
    """Fenêtre principale de l'application"""

    def __init__(self):
        super().__init__()
        self.current_user = None
        self.files_data = []
        self.file_processor = FileProcessor()
        self.selected_file_id = None
        self.selected_supplier_filter = None  # Filtre par fournisseur
        self.supplier_buttons = {}  # Dict pour stocker les boutons de fournisseurs
        self.is_loading = True  # Indicateur de chargement initial

        # Timer de rafraîchissement automatique
        self.refresh_timer = QTimer(self)
        self.refresh_timer.timeout.connect(self.auto_refresh_files)

        self.init_ui()

        # Configurer le rafraîchissement automatique selon les préférences
        self.setup_auto_refresh()

        # Afficher la boîte de dialogue de connexion
        self.show_login_dialog()

    def start_initial_load(self):
        """Démarre le chargement initial (appelé depuis main.py)"""
        # Forcer le rafraîchissement pour afficher le loader
        from PySide6.QtWidgets import QApplication
        QApplication.processEvents()

        # Lancer le chargement des données
        self.refresh_files_list()

        # Cacher le loader initial après le chargement complet
        self.is_loading = False
        if hasattr(self, 'initial_loader') and self.initial_loader.isVisible():
            self.initial_loader.hide()
            self.initial_spinner.timer.stop()  # Arrêter l'animation

    def resizeEvent(self, event):
        """Gère le redimensionnement de la fenêtre"""
        super().resizeEvent(event)

        # Repositionner le loader initial s'il est visible
        if hasattr(self, 'initial_loader') and self.initial_loader.isVisible():
            self.initial_loader.setGeometry(self.centralWidget().rect())

        # Réorganiser la grille des fournisseurs si elle existe et n'est pas en chargement
        if hasattr(self, 'supplier_buttons') and len(self.supplier_buttons) > 0 and not self.is_loading:
            self.load_suppliers_grid()

    def init_ui(self):
        """Initialise l'interface utilisateur"""
        self.setWindowTitle("Gestionnaire de Commandes Fournisseurs")

        # Forcer l'ouverture en plein écran sur l'écran principal
        from PySide6.QtWidgets import QApplication
        screen = QApplication.primaryScreen()
        if screen:
            geometry = screen.availableGeometry()
            self.setGeometry(geometry)
        self.showMaximized()  # Maximiser la fenêtre

        # Widget central
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # Layout principal
        main_layout = QVBoxLayout(central_widget)

        # Barre d'outils
        self.create_toolbar()

        # Titre de la section
        title_label = QLabel("📋 Fichiers FTP non commandés")
        title_label.setStyleSheet("font-size: 14pt; font-weight: bold; padding: 10px;")
        main_layout.addWidget(title_label)

        # Layout horizontal: Tableau + Grille de logos
        content_layout = QHBoxLayout()

        # Conteneur pour le tableau avec overlay de loading
        self.table_container = QFrame()
        self.table_container.setFrameStyle(QFrame.NoFrame)
        table_container_layout = QVBoxLayout(self.table_container)
        table_container_layout.setContentsMargins(0, 0, 0, 0)
        table_container_layout.setSpacing(0)

        # Tableau des fichiers (à gauche)
        self.files_table = QTableWidget()
        self.setup_files_table()
        table_container_layout.addWidget(self.files_table)

        content_layout.addWidget(self.table_container, 1)  # 33% de l'espace (1/3)

        # Grille de fournisseurs (à droite)
        suppliers_panel = self.create_suppliers_panel()
        content_layout.addWidget(suppliers_panel, 2)  # 66% de l'espace (2/3)

        main_layout.addLayout(content_layout)

        # Widget de chargement (remplace le tableau pendant le refresh)
        self.loading_widget = QWidget()
        loading_layout = QVBoxLayout(self.loading_widget)
        loading_layout.addStretch()

        self.loading_text = QLabel("🔄 Chargement en cours...")
        self.loading_text.setStyleSheet("font-size: 14pt; font-weight: bold; color: #2196F3;")
        self.loading_text.setAlignment(Qt.AlignCenter)
        loading_layout.addWidget(self.loading_text)

        self.progress_bar = QProgressBar()
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 2px solid #2196F3;
                border-radius: 5px;
                text-align: center;
                height: 30px;
                font-weight: bold;
            }
            QProgressBar::chunk {
                background-color: #2196F3;
            }
        """)
        loading_layout.addWidget(self.progress_bar)
        loading_layout.addStretch()

        self.loading_widget.setVisible(False)
        table_container_layout.addWidget(self.loading_widget)

        # Barre de statut
        self.statusBar = QStatusBar()
        self.setStatusBar(self.statusBar)
        self.statusBar.showMessage("Prêt")

        # Loader initial (overlay fullscreen pendant le premier chargement)
        self.initial_loader = QWidget(central_widget)
        self.initial_loader.setStyleSheet("""
            QWidget {
                background-color: rgba(255, 255, 255, 250);
            }
        """)
        initial_loader_layout = QVBoxLayout(self.initial_loader)
        initial_loader_layout.setAlignment(Qt.AlignCenter)

        # Logo dans le loader
        logo_path = get_resource_path("assets/logo/logo.png")
        if logo_path.exists():
            logo_label = QLabel()
            logo_pixmap = QPixmap(str(logo_path))
            # Redimensionner le logo à 128x128
            scaled_logo = logo_pixmap.scaled(128, 128, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            logo_label.setPixmap(scaled_logo)
            logo_label.setAlignment(Qt.AlignCenter)
            initial_loader_layout.addWidget(logo_label)

        # Titre dans le loader
        loader_title = QLabel("M-Jardin")
        loader_title.setStyleSheet("font-size: 32pt; font-weight: bold; color: #5abb44;")
        loader_title.setAlignment(Qt.AlignCenter)
        initial_loader_layout.addWidget(loader_title)

        # Importer le spinner du splash screen
        from app.ui.splash_screen import LoadingSpinner
        self.initial_spinner = LoadingSpinner()
        spinner_container = QWidget()
        spinner_layout = QVBoxLayout(spinner_container)
        spinner_layout.setAlignment(Qt.AlignCenter)
        spinner_layout.addWidget(self.initial_spinner)
        spinner_container.setStyleSheet("background: transparent;")
        initial_loader_layout.addWidget(spinner_container)

        loader_message = QLabel("Chargement de l'assistant de commande M-Jardin")
        loader_message.setStyleSheet("font-size: 12pt; color: #666;")
        loader_message.setAlignment(Qt.AlignCenter)
        initial_loader_layout.addWidget(loader_message)

        # Positionner le loader pour couvrir tout le central widget
        self.initial_loader.setGeometry(central_widget.rect())
        self.initial_loader.raise_()
        self.initial_loader.show()

    def _load_colored_svg_icon(self, svg_filename: str, color: str = "#5abb44") -> QIcon:
        """Charge un SVG et remplace sa couleur fill"""
        svg_path = get_resource_path(f"assets/icons/{svg_filename}")

        try:
            # Lire le fichier SVG
            with open(svg_path, 'r', encoding='utf-8') as f:
                svg_content = f.read()

            # Remplacer toutes les couleurs fill par la couleur souhaitée
            import re
            svg_content = re.sub(r'fill="[^"]*"', f'fill="{color}"', svg_content)
            svg_content = re.sub(r"fill='[^']*'", f"fill='{color}'", svg_content)

            # Créer un QPixmap à partir du SVG modifié
            svg_bytes = QByteArray(svg_content.encode('utf-8'))
            renderer = QSvgRenderer(svg_bytes)

            pixmap = QPixmap(32, 32)
            pixmap.fill(Qt.transparent)

            painter = QPainter(pixmap)
            renderer.render(painter)
            painter.end()

            return QIcon(pixmap)
        except Exception as e:
            logger.error(f"Erreur chargement icône SVG {svg_filename}: {e}")
            return QIcon()

    def create_toolbar(self):
        """Crée la barre d'outils"""
        toolbar = QToolBar("Barre d'outils principale")
        toolbar.setIconSize(QSize(32, 32))  # Taille des icônes augmentée
        toolbar.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)  # Texte à côté de l'icône

        # Style CSS pour augmenter la taille du menu
        toolbar.setStyleSheet("""
            QToolBar {
                spacing: 10px;
                padding: 8px;
            }
            QToolButton {
                font-size: 13pt;
                padding: 8px 12px;
                margin: 2px;
            }
        """)

        self.addToolBar(toolbar)

        # Action Rafraîchir
        refresh_action = QAction(self._load_colored_svg_icon("refresh.svg"), "Rafraîchir", self)
        refresh_action.triggered.connect(self.refresh_files_list)
        toolbar.addAction(refresh_action)

        toolbar.addSeparator()

        # Action Analyse des Commandes
        analysis_action = QAction(self._load_colored_svg_icon("analis.svg"), "Analyse Commandes", self)
        analysis_action.triggered.connect(self.open_order_analysis)
        toolbar.addAction(analysis_action)

        toolbar.addSeparator()

        # Action Gestion Fournisseurs
        suppliers_action = QAction(self._load_colored_svg_icon("fournisseur.svg"), "Gérer Fournisseurs", self)
        suppliers_action.triggered.connect(self.open_suppliers_manager)
        toolbar.addAction(suppliers_action)

        toolbar.addSeparator()

        # Action Réglages
        settings_action = QAction(self._load_colored_svg_icon("settings.svg"), "Réglages", self)
        settings_action.triggered.connect(self.open_settings)
        toolbar.addAction(settings_action)

        toolbar.addSeparator()

        # Action Déconnexion
        logout_action = QAction(self._load_colored_svg_icon("deconnect.svg"), "Déconnexion", self)
        logout_action.triggered.connect(self.logout)
        toolbar.addAction(logout_action)

    def create_suppliers_panel(self) -> QWidget:
        """Crée le panneau avec la grille des fournisseurs"""
        panel = QFrame()
        panel.setFrameStyle(QFrame.StyledPanel | QFrame.Raised)
        panel.setStyleSheet("""
            QFrame {
                background-color: #f9f9f9;
                border: 1px solid #ddd;
                border-radius: 5px;
                padding: 5px;
            }
        """)

        layout = QVBoxLayout(panel)

        # Titre
        title = QLabel("Fournisseurs")
        title.setStyleSheet("font-weight: bold; font-size: 12pt; padding: 5px;")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # Bouton "Tous"
        all_btn = QPushButton("Voir tous les fournisseurs")
        all_btn.setIcon(self._load_colored_svg_icon("all.svg", color="white"))
        all_btn.setIconSize(QSize(24, 24))
        all_btn.setStyleSheet("""
            QPushButton {
                padding: 8px 12px;
                background-color: #5abb44;
                color: white;
                border: none;
                border-radius: 4px;
                font-weight: bold;
                font-size: 11pt;
                text-align: center;
            }
            QPushButton:hover {
                background-color: #186108;
            }
        """)
        all_btn.clicked.connect(self.clear_supplier_filter)
        layout.addWidget(all_btn)

        # Scroll area pour la grille
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

        # Widget contenant la grille
        grid_widget = QWidget()
        self.suppliers_grid = QGridLayout(grid_widget)
        self.suppliers_grid.setSpacing(10)
        scroll.setWidget(grid_widget)

        layout.addWidget(scroll)

        # Zone en deux colonnes: statistiques à gauche, boutons d'action à droite
        bottom_zones = QWidget()
        bottom_zones_layout = QHBoxLayout(bottom_zones)
        bottom_zones_layout.setContentsMargins(0, 10, 0, 0)
        bottom_zones_layout.setSpacing(10)

        # Zone gauche: Statistiques (50%)
        stats_container = QFrame()
        stats_container.setStyleSheet("""
            QFrame {
                background-color: white;
                border: 2px solid #2196F3;
                border-radius: 8px;
                padding: 15px;
            }
        """)
        stats_layout = QVBoxLayout(stats_container)
        stats_layout.setSpacing(10)

        # Titre des statistiques
        stats_title = QLabel("📊 Statistiques des fichiers filtrés")
        stats_title.setStyleSheet("font-weight: bold; font-size: 11pt; color: #2196F3;")
        stats_title.setAlignment(Qt.AlignCenter)
        stats_layout.addWidget(stats_title)

        # Ligne 1: Nombre total de lignes
        self.stats_total_lines = QLabel("Nombre total de lignes : -")
        self.stats_total_lines.setStyleSheet("font-size: 10pt; padding: 5px;")
        self.stats_total_lines.setAlignment(Qt.AlignCenter)
        stats_layout.addWidget(self.stats_total_lines)

        # Ligne 2: Montant total
        self.stats_total_amount = QLabel("Montant total : -")
        self.stats_total_amount.setStyleSheet("font-size: 10pt; padding: 5px;")
        self.stats_total_amount.setAlignment(Qt.AlignCenter)
        stats_layout.addWidget(self.stats_total_amount)

        bottom_zones_layout.addWidget(stats_container, 1)

        # Zone droite: Boutons d'action (50%)
        actions_container = QFrame()
        actions_container.setStyleSheet("""
            QFrame {
                background-color: white;
                border: 2px solid #4CAF50;
                border-radius: 8px;
                padding: 15px;
            }
        """)
        actions_layout = QVBoxLayout(actions_container)
        actions_layout.setSpacing(10)

        # Titre de la zone actions
        actions_title = QLabel("⚡ Actions rapides")
        actions_title.setStyleSheet("font-weight: bold; font-size: 11pt; color: #4CAF50;")
        actions_title.setAlignment(Qt.AlignCenter)
        actions_layout.addWidget(actions_title)

        # Container horizontal pour les trois boutons
        buttons_container = QWidget()
        buttons_layout = QHBoxLayout(buttons_container)
        buttons_layout.setSpacing(15)
        buttons_layout.setContentsMargins(10, 10, 10, 10)

        # Style commun pour les boutons d'action
        action_button_style = """
            QPushButton {
                background-color: #5abb44;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 15px 10px;
                font-weight: bold;
                font-size: 10pt;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #186108;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
        """

        # Bouton OUVRIR
        open_btn = QPushButton("OUVRIR")
        open_btn.setIcon(self._load_colored_svg_icon("open-file.svg", color="white"))
        open_btn.setIconSize(QSize(48, 48))
        open_btn.setStyleSheet(action_button_style)
        open_btn.setToolTip("Ouvrir le fichier sélectionné")
        open_btn.setCursor(Qt.PointingHandCursor)  # Curseur pointeur
        open_btn.clicked.connect(self.open_file)
        open_btn.setEnabled(False)  # Désactivé par défaut
        buttons_layout.addWidget(open_btn)
        self.open_btn = open_btn

        # Bouton WEB
        web_btn = QPushButton("WEB")
        web_btn.setIcon(self._load_colored_svg_icon("web.svg", color="white"))
        web_btn.setIconSize(QSize(48, 48))
        web_btn.setStyleSheet(action_button_style)
        web_btn.setToolTip("Accéder au site web du fournisseur")
        web_btn.setCursor(Qt.PointingHandCursor)  # Curseur pointeur
        web_btn.clicked.connect(self.open_supplier_website)
        web_btn.setEnabled(False)  # Désactivé par défaut
        buttons_layout.addWidget(web_btn)
        self.web_btn = web_btn

        # Bouton EXPORTER
        export_btn = QPushButton("EXPORTER")
        export_btn.setIcon(self._load_colored_svg_icon("export.svg", color="white"))
        export_btn.setIconSize(QSize(48, 48))
        export_btn.setStyleSheet(action_button_style)
        export_btn.setToolTip("Exporter les données")
        export_btn.setCursor(Qt.PointingHandCursor)  # Curseur pointeur
        export_btn.clicked.connect(self.export_data)
        export_btn.setEnabled(False)  # Désactivé par défaut
        buttons_layout.addWidget(export_btn)
        self.export_btn = export_btn

        # Bouton IMPRIMER
        print_btn = QPushButton("IMPRIMER")
        print_btn.setIcon(self._load_colored_svg_icon("print.svg", color="white"))
        print_btn.setIconSize(QSize(48, 48))
        print_btn.setStyleSheet(action_button_style)
        print_btn.setToolTip("Imprimer la sélection")
        print_btn.setCursor(Qt.PointingHandCursor)  # Curseur pointeur
        print_btn.clicked.connect(self.print_selection)
        print_btn.setEnabled(False)  # Désactivé par défaut
        buttons_layout.addWidget(print_btn)
        self.print_btn = print_btn

        # Bouton ARCHIVER
        archive_btn = QPushButton("ARCHIVER")
        archive_btn.setIcon(self._load_colored_svg_icon("archive.svg", color="white"))
        archive_btn.setIconSize(QSize(48, 48))
        archive_btn.setStyleSheet(action_button_style)
        archive_btn.setToolTip("Archiver le fichier sélectionné")
        archive_btn.setCursor(Qt.PointingHandCursor)  # Curseur pointeur
        archive_btn.clicked.connect(self.archive_file)
        archive_btn.setEnabled(False)  # Désactivé par défaut
        buttons_layout.addWidget(archive_btn)
        self.new_archive_btn = archive_btn

        actions_layout.addWidget(buttons_container)

        bottom_zones_layout.addWidget(actions_container, 1)

        layout.addWidget(bottom_zones)

        return panel

    def create_filters_section(self) -> QHBoxLayout:
        """Crée la section des filtres"""
        layout = QHBoxLayout()

        # Filtre par statut
        layout.addWidget(QLabel("Statut:"))
        self.status_filter = QComboBox()
        self.status_filter.addItems(["Tous", "À traiter", "En cours", "Terminé", "Erreur"])
        self.status_filter.currentTextChanged.connect(self.apply_filters)
        layout.addWidget(self.status_filter)

        # Filtre par date
        layout.addWidget(QLabel("Date:"))
        self.date_filter = QDateEdit()
        self.date_filter.setCalendarPopup(True)
        self.date_filter.setDate(QDate.currentDate())
        self.date_filter.dateChanged.connect(self.apply_filters)
        layout.addWidget(self.date_filter)

        # Bouton afficher tout
        show_all_btn = QPushButton("Afficher tout")
        show_all_btn.clicked.connect(self.show_all_files)
        layout.addWidget(show_all_btn)

        layout.addStretch()

        return layout

    def setup_files_table(self):
        """Configure le tableau des fichiers FTP"""
        self.files_table.setColumnCount(4)
        self.files_table.setHorizontalHeaderLabels([
            "Nom du fichier", "Date", "Chemin complet", "☑"
        ])

        # Masquer la colonne chemin complet
        self.files_table.setColumnHidden(2, True)

        # Masquer les numéros de ligne (en-têtes verticaux)
        self.files_table.verticalHeader().setVisible(False)

        # Configurer les largeurs de colonnes
        from PySide6.QtWidgets import QHeaderView
        header = self.files_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Stretch)  # Nom du fichier (stretch pour remplir)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # Date
        header.setSectionResizeMode(3, QHeaderView.Fixed)  # Checkbox (largeur fixe)
        header.resizeSection(3, 40)  # 40 pixels pour la checkbox

        # Appliquer le delegate personnalisé pour empêcher le style de sélection Qt
        self.files_table.setItemDelegate(NoSelectionDelegate())

        # Sélection par ligne
        self.files_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.files_table.setSelectionMode(QTableWidget.SingleSelection)

        # Connexion du signal de sélection
        self.files_table.itemSelectionChanged.connect(self.on_file_selected)

    # Ancienne section de boutons d'action (remplacée par les boutons dans "Actions rapides")
    # def create_actions_section(self) -> QHBoxLayout:
    #     """Crée la section des boutons d'action"""
    #     layout = QHBoxLayout()
    #
    #     self.download_btn = QPushButton("📥 Télécharger")
    #     self.download_btn.clicked.connect(self.download_file)
    #     self.download_btn.setEnabled(False)
    #     self.download_btn.setToolTip("Télécharger le fichier localement")
    #     layout.addWidget(self.download_btn)
    #
    #     self.transform_btn = QPushButton("⚙️ Transformer")
    #     self.transform_btn.clicked.connect(self.transform_file)
    #     self.transform_btn.setEnabled(False)
    #     self.transform_btn.setToolTip("Télécharger et transformer le fichier")
    #     layout.addWidget(self.transform_btn)
    #
    #     self.archive_btn = QPushButton("📦 Archiver (→ old)")
    #     self.archive_btn.clicked.connect(self.archive_file)
    #     self.archive_btn.setEnabled(False)
    #     self.archive_btn.setToolTip("Déplace le fichier vers le dossier 'old' sur le serveur FTP")
    #     layout.addWidget(self.archive_btn)
    #
    #     layout.addStretch()
    #
    #     return layout

    def show_login_dialog(self):
        """Affiche la boîte de dialogue de connexion"""
        from app.ui.login_dialog import LoginDialog

        dialog = LoginDialog(self)
        if dialog.exec() == QDialog.Accepted:
            self.current_user = supabase_client.get_current_user()
            if self.current_user:
                user_email = getattr(self.current_user, 'email', 'Utilisateur')
                logger.info(f"Utilisateur connecté: {user_email}")
                self.statusBar.showMessage(f"Connecté: {user_email}")
                self.refresh_files_list()
            else:
                QMessageBox.critical(self, "Erreur", "Impossible de récupérer les informations utilisateur")
                sys.exit(1)
        else:
            sys.exit(0)

    def refresh_files_list(self):
        """Rafraîchit la liste des fichiers depuis le serveur FTP"""
        logger.debug("Rafraîchissement de la liste des fichiers FTP")

        # Vider le cache des statistiques
        if hasattr(self, '_file_stats_cache'):
            self._file_stats_cache.clear()

        # Si le loader initial est visible, le garder au-dessus
        if hasattr(self, 'initial_loader') and self.initial_loader.isVisible():
            self.initial_loader.raise_()

        # Cacher le tableau, afficher le widget de chargement (uniquement si pas de loader initial)
        self.files_table.setVisible(False)
        if not (hasattr(self, 'initial_loader') and self.initial_loader.isVisible()):
            self.loading_widget.setVisible(True)
        self.progress_bar.setValue(0)
        self.loading_text.setText("🔄 Connexion au serveur FTP...")

        # Désactiver les boutons d'action pendant le chargement
        if hasattr(self, 'print_btn'):
            self.print_btn.setEnabled(False)
        if hasattr(self, 'export_btn'):
            self.export_btn.setEnabled(False)
        if hasattr(self, 'new_archive_btn'):
            self.new_archive_btn.setEnabled(False)
        if hasattr(self, 'open_btn'):
            self.open_btn.setEnabled(False)
        if hasattr(self, 'web_btn'):
            self.web_btn.setEnabled(False)

        # Forcer le rafraîchissement de l'interface et permettre au spinner de tourner
        from PySide6.QtWidgets import QApplication
        QApplication.processEvents()

        try:
            # Récupérer les paramètres FTP depuis .env
            ftp_host = os.getenv("FTP_HOST")
            ftp_port = int(os.getenv("FTP_PORT", 22))
            ftp_user = os.getenv("FTP_USERNAME")
            ftp_pass = os.getenv("FTP_PASSWORD")
            ftp_path = os.getenv("FTP_REMOTE_PATH", "/home/mjard_ep43/export-cdes-fournisseurs")

            if not all([ftp_host, ftp_user, ftp_pass]):
                raise ValueError("Configuration FTP incomplète dans .env")

            self.statusBar.showMessage("🔄 Connexion au serveur FTP...")
            self.progress_bar.setValue(10)
            if hasattr(self, 'initial_loader') and self.initial_loader.isVisible():
                self.initial_loader.raise_()
            QApplication.processEvents()  # Permettre au spinner de tourner

            # Se connecter au FTP
            fetcher = FTPFetcher(ftp_host, ftp_port, ftp_user, ftp_pass, use_sftp=True)
            fetcher.connect()

            self.loading_text.setText("📋 Récupération de la liste des fichiers...")
            self.progress_bar.setValue(30)
            if hasattr(self, 'initial_loader') and self.initial_loader.isVisible():
                self.initial_loader.raise_()
            QApplication.processEvents()  # Permettre au spinner de tourner
            self.statusBar.showMessage("📋 Récupération de la liste des fichiers...")

            # Lister les fichiers (exclure le dossier 'old')
            self.files_data = fetcher.list_files(ftp_path, exclude_dirs=['old'])

            # Déconnexion
            fetcher.disconnect()

            self.loading_text.setText(f"📊 Analyse de {len(self.files_data)} fichier(s) en cours...")
            if hasattr(self, 'initial_loader') and self.initial_loader.isVisible():
                self.initial_loader.raise_()
            QApplication.processEvents()  # Permettre au spinner de tourner
            self.progress_bar.setValue(60)
            self.loading_widget.repaint()
            self.statusBar.showMessage(f"📊 Analyse de {len(self.files_data)} fichier(s) en cours...")

            # Charger les fournisseurs dans la grille
            self.load_suppliers_grid()
            self.progress_bar.setValue(80)
            if hasattr(self, 'initial_loader') and self.initial_loader.isVisible():
                self.initial_loader.raise_()
            QApplication.processEvents()  # Permettre au spinner de tourner

            # Afficher dans le tableau
            self.populate_ftp_table(self.files_data)

            self.progress_bar.setValue(100)
            if hasattr(self, 'initial_loader') and self.initial_loader.isVisible():
                self.initial_loader.raise_()
            QApplication.processEvents()  # Permettre au spinner de tourner
            self.statusBar.showMessage(f"✅ {len(self.files_data)} fichier(s) non commandé(s) trouvé(s)", 5000)

        except Exception as e:
            logger.error(f"Erreur lors du rafraîchissement FTP: {e}")
            QMessageBox.warning(self, "Erreur", f"Impossible de rafraîchir la liste FTP:\n{str(e)}")
            self.statusBar.showMessage("❌ Erreur de connexion FTP")

        finally:
            # Cacher le loader, réafficher le tableau
            self.loading_widget.setVisible(False)
            self.files_table.setVisible(True)

    def load_suppliers_grid(self):
        """Charge les fournisseurs actifs dans la grille"""
        try:
            # Vider la grille
            while self.suppliers_grid.count():
                item = self.suppliers_grid.takeAt(0)
                if item.widget():
                    item.widget().deleteLater()

            self.supplier_buttons.clear()

            # Récupérer les fournisseurs actifs depuis la BDD
            response = supabase_client.client.table('suppliers').select('*').eq('active', True).order('name').execute()
            suppliers = response.data

            # Calculer le nombre de colonnes en fonction de la largeur disponible
            # Largeur estimée d'un widget: 140px (logo) + 20px (padding) + 10px (spacing) = 170px
            panel_width = self.suppliers_grid.parentWidget().width() if self.suppliers_grid.parentWidget() else 800
            widget_width = 155  # Largeur réduite de 10%: 140px -> 126px + padding
            max_cols = max(2, panel_width // widget_width)  # Minimum 2 colonnes

            row = 0
            col = 0

            for supplier in suppliers:
                supplier_id = supplier.get('id')
                supplier_name = supplier.get('name', '')
                logo_url = supplier.get('logo_url', '')
                file_filter_slug = supplier.get('file_filter_slug', '')

                # Créer un widget pour le fournisseur
                supplier_widget = QFrame()
                supplier_widget.setObjectName("supplierCard")  # Nom unique pour cibler ce frame spécifiquement
                supplier_widget.setFrameStyle(QFrame.Box | QFrame.Raised)
                supplier_widget.setMaximumWidth(146)  # Largeur max: 126px logo + 20px padding
                supplier_widget.setStyleSheet("""
                    QFrame#supplierCard {
                        background-color: white;
                        border: 2px solid #ddd;
                        border-radius: 8px;
                        padding: 10px;
                    }
                    QFrame#supplierCard:hover {
                        border-color: #2196F3;
                    }
                """)
                supplier_widget.setCursor(Qt.PointingHandCursor)

                widget_layout = QVBoxLayout(supplier_widget)
                widget_layout.setSpacing(5)
                widget_layout.setContentsMargins(5, 5, 5, 5)

                # Logo avec fond blanc toujours
                logo_label = QLabel()
                logo_label.setAlignment(Qt.AlignCenter)
                logo_label.setFixedHeight(90)  # Réduit de 10%: 100px -> 90px
                logo_label.setMaximumWidth(126)  # Réduit de 10%: 140px -> 126px
                logo_label.setScaledContents(False)
                logo_label.setStyleSheet("background-color: white;")  # Fond blanc permanent

                if logo_url:
                    self._load_supplier_logo(logo_label, logo_url, 104)  # Réduit de 10%: 115px -> 104px
                else:
                    # Placeholder si pas de logo
                    logo_label.setText("📦")
                    logo_label.setStyleSheet("font-size: 72px; background-color: white;")

                # Centrer le logo dans le layout
                logo_container = QWidget()
                logo_container_layout = QHBoxLayout(logo_container)
                logo_container_layout.setContentsMargins(0, 0, 0, 0)
                logo_container_layout.addStretch()
                logo_container_layout.addWidget(logo_label)
                logo_container_layout.addStretch()
                logo_container.setStyleSheet("background-color: transparent;")
                widget_layout.addWidget(logo_container)

                # Nom - hauteur adaptée au contenu
                name_label = QLabel(supplier_name)
                name_label.setAlignment(Qt.AlignCenter)
                name_label.setWordWrap(True)
                name_label.setMaximumWidth(126)  # Même largeur que le logo
                name_label.setStyleSheet("font-weight: bold; font-size: 10pt; padding: 2px;")
                name_label.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Minimum)  # Hauteur minimale adaptée
                widget_layout.addWidget(name_label)

                # Empêcher l'étirement vertical
                widget_layout.addStretch()
                supplier_widget.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)

                # Connecter le clic
                supplier_widget.mousePressEvent = lambda event, slug=file_filter_slug: self.filter_by_supplier(slug)

                # Ajouter à la grille
                self.suppliers_grid.addWidget(supplier_widget, row, col)

                # Stocker la référence
                self.supplier_buttons[file_filter_slug] = supplier_widget

                # Passer à la colonne/ligne suivante
                col += 1
                if col >= max_cols:
                    col = 0
                    row += 1

            logger.info(f"{len(suppliers)} fournisseur(s) actif(s) chargé(s) dans la grille")

        except Exception as e:
            logger.error(f"Erreur chargement grille fournisseurs: {e}")

    def _load_supplier_logo(self, label: QLabel, url: str, size: int):
        """Charge et affiche un logo de fournisseur"""
        try:
            import requests

            response = requests.get(url, timeout=5)
            response.raise_for_status()

            pixmap = QPixmap()
            pixmap.loadFromData(response.content)

            if not pixmap.isNull():
                scaled_pixmap = pixmap.scaled(
                    size, size,
                    Qt.KeepAspectRatio,
                    Qt.SmoothTransformation
                )
                label.setPixmap(scaled_pixmap)

        except Exception as e:
            logger.error(f"Erreur chargement logo fournisseur {url}: {e}")

    def filter_by_supplier(self, file_filter_slug: str):
        """Filtre les fichiers par fournisseur"""
        self.selected_supplier_filter = file_filter_slug
        logger.info(f"Filtre appliqué: {file_filter_slug}")

        # Mettre à jour le style des boutons
        for slug, widget in self.supplier_buttons.items():
            if slug == file_filter_slug:
                # Actif (vert) - uniquement le cadre extérieur
                widget.setStyleSheet("""
                    QFrame#supplierCard {
                        background-color: #5abb44;
                        border: 3px solid #5abb44;
                        border-radius: 8px;
                        padding: 10px;
                    }
                    QLabel {
                        color: white;
                        background-color: transparent;
                    }
                """)
            else:
                # Inactif (grisé avec hover bleu sur la bordure)
                widget.setStyleSheet("""
                    QFrame#supplierCard {
                        background-color: #f5f5f5;
                        border: 2px solid #ddd;
                        border-radius: 8px;
                        padding: 10px;
                        opacity: 0.5;
                    }
                    QFrame#supplierCard:hover {
                        border-color: #2196F3;
                    }
                """)

        # Appliquer le filtre au tableau
        self.apply_filters()

    def clear_supplier_filter(self):
        """Retire le filtre fournisseur"""
        self.selected_supplier_filter = None
        logger.info("Filtre fournisseur retiré")

        # Remettre tous les boutons en style normal
        for widget in self.supplier_buttons.values():
            widget.setStyleSheet("""
                QFrame#supplierCard {
                    background-color: white;
                    border: 2px solid #ddd;
                    border-radius: 8px;
                    padding: 10px;
                }
                QFrame#supplierCard:hover {
                    border-color: #2196F3;
                }
                QLabel {
                    background-color: transparent;
                }
            """)

        # Appliquer le filtre (vide) au tableau
        self.apply_filters()

    def apply_filters(self):
        """Applique les filtres à la liste des fichiers FTP"""
        # Pour l'instant, on affiche simplement tous les fichiers
        # Les filtres peuvent être ajoutés plus tard si nécessaire
        self.populate_ftp_table(self.files_data)

    def show_all_files(self):
        """Affiche tous les fichiers sans filtre de date"""
        self.show_all_mode = True
        self.apply_filters()

    def populate_ftp_table(self, files: list):
        """Remplit le tableau avec les fichiers FTP"""
        self.files_table.setRowCount(0)

        # Filtrer par fournisseur si un filtre est actif
        if self.selected_supplier_filter:
            # Récupérer les patterns du fournisseur depuis la base de données
            try:
                response = supabase_client.client.table('suppliers').select('file_patterns').eq('file_filter_slug', self.selected_supplier_filter).execute()
                if response.data and len(response.data) > 0:
                    patterns = response.data[0].get('file_patterns', [])

                    if patterns:
                        # Filtrer les fichiers selon les patterns
                        import fnmatch
                        filtered_files = []
                        for file_info in files:
                            filename = file_info.get('filename', '')
                            # Vérifier si le fichier correspond à au moins un pattern
                            for pattern in patterns:
                                if fnmatch.fnmatch(filename, pattern):
                                    filtered_files.append(file_info)
                                    break  # Pas besoin de vérifier les autres patterns

                        files = filtered_files
                        logger.debug(f"{len(files)} fichier(s) après filtrage par patterns {patterns}")
                    else:
                        # Pas de patterns configurés, utiliser le slug comme fallback
                        files = [f for f in files if f.get('filename', '').startswith(self.selected_supplier_filter)]
                        logger.debug(f"{len(files)} fichier(s) après filtrage par slug '{self.selected_supplier_filter}'")
                else:
                    logger.warning(f"Fournisseur '{self.selected_supplier_filter}' introuvable")
                    files = []
            except Exception as e:
                logger.error(f"Erreur lors de la récupération des patterns: {e}")
                # Fallback sur le slug
                files = [f for f in files if f.get('filename', '').startswith(self.selected_supplier_filter)]
                logger.debug(f"{len(files)} fichier(s) après filtrage par slug '{self.selected_supplier_filter}' (fallback)")

        # Trier les fichiers par ordre alphabétique selon le nom
        files_sorted = sorted(files, key=lambda f: f.get('filename', '').lower())

        # Récupérer le chemin FTP de base
        ftp_path = os.getenv("FTP_REMOTE_PATH", "/home/mjard_ep43/export-cdes-fournisseurs")

        for file_info in files_sorted:
            row = self.files_table.rowCount()
            self.files_table.insertRow(row)

            # Colonne 0: Nom du fichier (avec padding)
            filename = file_info.get('filename', '')
            filename_item = QTableWidgetItem(f"  {filename}")  # Padding à gauche
            self.files_table.setItem(row, 0, filename_item)

            # Colonne 1: Date
            modified = file_info.get('modified')
            if modified:
                date_str = modified.strftime("%d/%m/%Y")
            else:
                date_str = "-"
            date_item = QTableWidgetItem(date_str)
            date_item.setTextAlignment(Qt.AlignCenter)
            self.files_table.setItem(row, 1, date_item)

            # Colonne 2: Chemin complet (caché)
            full_path = f"{ftp_path}/{filename}".replace('//', '/')
            self.files_table.setItem(row, 2, QTableWidgetItem(full_path))

            # Colonne 3: Checkbox (dernière colonne)
            checkbox_widget = QWidget()
            checkbox_layout = QHBoxLayout(checkbox_widget)
            checkbox = QCheckBox()
            is_checked = bool(self.selected_supplier_filter)
            checkbox.setChecked(is_checked)  # Coché si filtre actif
            checkbox.setProperty('row', row)  # Stocker le numéro de ligne dans la checkbox
            checkbox.stateChanged.connect(self.on_checkbox_changed)  # Mettre à jour les stats et le background
            checkbox_layout.addWidget(checkbox)
            checkbox_layout.setAlignment(Qt.AlignCenter)
            checkbox_layout.setContentsMargins(0, 0, 0, 0)
            self.files_table.setCellWidget(row, 3, checkbox_widget)

        # Si un filtre fournisseur est actif et qu'il y a des fichiers, sélectionner automatiquement le premier
        if self.selected_supplier_filter and self.files_table.rowCount() > 0:
            logger.debug(f"Sélection automatique de la ligne 0 (total: {self.files_table.rowCount()} lignes)")
            self.files_table.selectRow(0)
            # Le signal itemSelectionChanged déclenchera automatiquement on_file_selected()
            # Forcer l'appel manuel au cas où le signal ne se déclenche pas
            from PySide6.QtCore import QTimer
            QTimer.singleShot(100, self.on_file_selected)

        # Appliquer le background à toutes les lignes cochées avec un délai
        # pour laisser Qt finir ses opérations de style de sélection
        if self.selected_supplier_filter:
            from PySide6.QtCore import QTimer
            QTimer.singleShot(50, lambda: self.apply_all_backgrounds())

        # Mettre à jour les statistiques
        self.update_file_statistics()

    def apply_all_backgrounds(self):
        """Applique le background à toutes les lignes cochées"""
        for row in range(self.files_table.rowCount()):
            checkbox_widget = self.files_table.cellWidget(row, 3)
            if checkbox_widget:
                checkbox = checkbox_widget.findChild(QCheckBox)
                if checkbox and checkbox.isChecked():
                    self.set_row_background(row, True)

    def set_row_background(self, row: int, checked: bool):
        """Applique ou retire le background vert sur une ligne en fonction de l'état de la checkbox"""
        from PySide6.QtGui import QBrush, QColor

        logger.debug(f"set_row_background: row={row}, checked={checked}")

        # Déterminer la couleur de fond
        if checked:
            bg_color = QColor("#9df589")  # Vert
            logger.debug(f"  -> Application du fond VERT sur ligne {row}")
        else:
            bg_color = QColor(Qt.white)  # Blanc par défaut de Qt
            logger.debug(f"  -> Application du fond BLANC sur ligne {row}")

        # Appliquer le background à toutes les colonnes de la ligne
        for col in range(self.files_table.columnCount()):
            if not self.files_table.isColumnHidden(col):
                item = self.files_table.item(row, col)
                if item:
                    # Appliquer le background avec setData pour forcer
                    brush = QBrush(bg_color)
                    item.setBackground(brush)
                    item.setData(Qt.BackgroundRole, brush)

                # Pour les widgets (comme la checkbox en colonne 3)
                widget = self.files_table.cellWidget(row, col)
                if widget:
                    # Appliquer le stylesheet
                    if checked:
                        widget.setStyleSheet("QWidget { background-color: #9df589; }")
                    else:
                        widget.setStyleSheet("")  # Retirer le style

        # Forcer un rafraîchissement immédiat et complet
        self.files_table.viewport().update()
        self.files_table.repaint()

        logger.debug(f"  -> Rafraîchissement forcé pour ligne {row}")

    def on_checkbox_changed(self, state):
        """Appelé quand l'état d'une checkbox change"""
        # Récupérer la checkbox qui a émis le signal
        checkbox = self.sender()
        if checkbox:
            # Récupérer le numéro de ligne stocké dans la propriété
            row = checkbox.property('row')
            if row is not None:
                # Mettre à jour le background de la ligne
                # Utiliser bool(state) car state est un int: 0=unchecked, 2=checked
                is_checked = bool(state)
                logger.debug(f"on_checkbox_changed: row={row}, state={state}, is_checked={is_checked}")
                self.set_row_background(row, is_checked)

        # Mettre à jour les statistiques
        self.update_file_statistics()

    def get_checked_files(self):
        """Retourne la liste des fichiers cochés dans le tableau"""
        checked_files = []
        for row in range(self.files_table.rowCount()):
            # Récupérer le widget checkbox de la dernière colonne (colonne 3)
            checkbox_widget = self.files_table.cellWidget(row, 3)
            if checkbox_widget:
                checkbox = checkbox_widget.findChild(QCheckBox)
                if checkbox and checkbox.isChecked():
                    # Récupérer les informations du fichier
                    filename = self.files_table.item(row, 0).text().strip()  # Colonne 0: Nom (strip pour enlever le padding)
                    full_path = self.files_table.item(row, 2).text()  # Colonne 2: Chemin complet
                    checked_files.append({'filename': filename, 'full_path': full_path})
        return checked_files

    def update_file_statistics(self):
        """Met à jour les statistiques affichées (nb lignes et montant total) - uniquement pour les fichiers cochés"""
        # Récupérer les fichiers cochés
        checked_files = self.get_checked_files()

        if not checked_files:
            # Aucun fichier coché -> stats à zéro
            self.stats_total_lines.setText("Nombre total de lignes : 0")
            self.stats_total_amount.setText("Montant total : 0.00 €")
            return

        # Initialiser le cache si nécessaire
        if not hasattr(self, '_file_stats_cache'):
            self._file_stats_cache = {}

        total_lines = 0
        total_amount = 0.0

        # Connexion FTP pour télécharger et analyser les fichiers
        try:
            ftp_host = os.getenv("FTP_HOST")
            ftp_port = int(os.getenv("FTP_PORT", 22))
            ftp_user = os.getenv("FTP_USERNAME")
            ftp_pass = os.getenv("FTP_PASSWORD")

            fetcher = FTPFetcher(ftp_host, ftp_port, ftp_user, ftp_pass, use_sftp=True)
            if not fetcher.connect():
                logger.error("Impossible de se connecter au FTP pour calculer les statistiques")
                self.stats_total_lines.setText("Nombre total de lignes : Erreur connexion FTP")
                self.stats_total_amount.setText("Montant total : Erreur connexion FTP")
                return
        except Exception as e:
            logger.error(f"Erreur connexion FTP pour statistiques: {e}")
            self.stats_total_lines.setText("Nombre total de lignes : Erreur")
            self.stats_total_amount.setText("Montant total : Erreur")
            return

        # Analyser chaque fichier coché
        for file_info in checked_files:
            filename = file_info['filename']
            full_path = file_info['full_path']

            # Vérifier le cache
            if filename in self._file_stats_cache:
                cached_stats = self._file_stats_cache[filename]
                total_lines += cached_stats['lines']
                total_amount += cached_stats['amount']
                logger.debug(f"Utilisation du cache pour {filename}: {cached_stats['lines']} lignes, {cached_stats['amount']:.2f} €")
                continue

            # Télécharger et analyser le fichier
            tmp_path = None
            try:
                with tempfile.NamedTemporaryFile(mode='w+b', delete=False, suffix='.csv') as tmp_file:
                    tmp_path = tmp_file.name

                # Télécharger le fichier
                fetcher.download_file(full_path, tmp_path)

                # Analyser avec pandas
                df = None
                for encoding in ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252']:
                    try:
                        df = pd.read_csv(tmp_path, encoding=encoding, sep=';', on_bad_lines='skip', header=None)
                        break
                    except UnicodeDecodeError:
                        continue

                if df is None:
                    logger.error(f"Impossible de décoder {filename}")
                    continue

                # Compter les lignes
                line_count = len(df)
                total_lines += line_count

                # Calculer la somme de la colonne D (index 3)
                file_amount = 0.0
                if len(df.columns) >= 4:
                    column_d = df.iloc[:, 3]
                    column_d_numeric = pd.to_numeric(column_d, errors='coerce')
                    file_amount = column_d_numeric.sum()
                    total_amount += file_amount

                # Mettre en cache
                self._file_stats_cache[filename] = {
                    'lines': line_count,
                    'amount': file_amount
                }
                logger.debug(f"Analyse de {filename}: {line_count} lignes, {file_amount:.2f} € (mis en cache)")

            except Exception as e:
                logger.error(f"Erreur analyse {filename}: {e}")
            finally:
                if tmp_path and os.path.exists(tmp_path):
                    try:
                        os.unlink(tmp_path)
                    except:
                        pass

        # Déconnexion FTP
        fetcher.disconnect()

        # Mettre à jour les labels
        self.stats_total_lines.setText(f"Nombre total de lignes : {total_lines:,}".replace(',', ' '))
        self.stats_total_amount.setText(f"Montant total : {total_amount:,.2f} €".replace(',', ' '))

    def extract_supplier_from_filename(self, filename: str) -> str:
        """Extrait le nom du fournisseur depuis le nom du fichier"""
        # Format attendu: Fournisseur-DD-MM-YY.csv
        try:
            # Enlever l'extension
            name_without_ext = filename.rsplit('.', 1)[0]
            # Prendre la partie avant le premier tiret
            supplier = name_without_ext.split('-')[0]
            return supplier
        except:
            return "Inconnu"

    @Slot()
    def on_file_selected(self):
        """Gère la sélection d'un fichier FTP"""
        selected_rows = self.files_table.selectedItems()
        if not selected_rows:
            logger.debug("Aucun fichier sélectionné - désactivation des boutons")
            self.selected_file_id = None
            if hasattr(self, 'print_btn'):
                self.print_btn.setEnabled(False)
            if hasattr(self, 'export_btn'):
                self.export_btn.setEnabled(False)
            if hasattr(self, 'new_archive_btn'):
                self.new_archive_btn.setEnabled(False)
            if hasattr(self, 'open_btn'):
                self.open_btn.setEnabled(False)
            if hasattr(self, 'web_btn'):
                self.web_btn.setEnabled(False)
            return

        # Récupérer le chemin du fichier sélectionné (colonne 2 - chemin complet caché)
        row = selected_rows[0].row()
        self.selected_file_id = self.files_table.item(row, 2).text()
        logger.debug(f"Fichier sélectionné: {self.selected_file_id} - activation des boutons")

        # Activer tous les boutons quand un fichier est sélectionné
        if hasattr(self, 'print_btn'):
            self.print_btn.setEnabled(True)
        if hasattr(self, 'export_btn'):
            self.export_btn.setEnabled(True)
        if hasattr(self, 'new_archive_btn'):
            self.new_archive_btn.setEnabled(True)
        if hasattr(self, 'open_btn'):
            self.open_btn.setEnabled(True)
        if hasattr(self, 'web_btn'):
            self.web_btn.setEnabled(True)

    def download_file(self):
        """Télécharge le fichier sélectionné depuis le FTP"""
        if not self.selected_file_id:
            return

        try:
            # Demander où sauvegarder
            filename = Path(self.selected_file_id).name
            save_path, _ = QFileDialog.getSaveFileName(
                self,
                "Enregistrer le fichier",
                filename,
                "Tous les fichiers (*.*)"
            )

            if not save_path:
                return

            self.statusBar.showMessage("Téléchargement en cours...")

            # Connexion FTP
            ftp_host = os.getenv("FTP_HOST")
            ftp_port = int(os.getenv("FTP_PORT", 22))
            ftp_user = os.getenv("FTP_USERNAME")
            ftp_pass = os.getenv("FTP_PASSWORD")

            fetcher = FTPFetcher(ftp_host, ftp_port, ftp_user, ftp_pass, use_sftp=True)
            fetcher.connect()

            # Télécharger le fichier
            fetcher.download_file(self.selected_file_id, save_path)

            fetcher.disconnect()

            self.statusBar.showMessage(f"✅ Fichier téléchargé: {save_path}", 5000)
            QMessageBox.information(self, "Succès", f"Fichier téléchargé:\n{save_path}")

        except Exception as e:
            logger.error(f"Erreur téléchargement: {e}")
            QMessageBox.warning(self, "Erreur", f"Erreur lors du téléchargement:\n{str(e)}")
            self.statusBar.showMessage("❌ Erreur de téléchargement")

    def lock_file(self):
        """Verrouille le fichier sélectionné"""
        if not self.selected_file_id:
            return

        user_id = getattr(self.current_user, 'id', None)
        success = supabase_client.lock_file(self.selected_file_id, user_id)

        if success:
            QMessageBox.information(self, "Succès", "Fichier verrouillé avec succès")
            self.refresh_files_list()
        else:
            QMessageBox.warning(self, "Erreur", "Impossible de verrouiller le fichier")

    def unlock_file(self):
        """Déverrouille le fichier sélectionné"""
        if not self.selected_file_id:
            return

        user_id = getattr(self.current_user, 'id', None)
        success = supabase_client.unlock_file(self.selected_file_id, user_id)

        if success:
            QMessageBox.information(self, "Succès", "Fichier déverrouillé")
            self.refresh_files_list()
        else:
            QMessageBox.warning(self, "Erreur", "Impossible de déverrouiller le fichier")

    def transform_file(self):
        """Transforme le fichier sélectionné"""
        QMessageBox.information(self, "Info", "Fonctionnalité de transformation à implémenter avec vos règles spécifiques")

    def merge_files(self):
        """Regroupe plusieurs fichiers"""
        QMessageBox.information(self, "Info", "Fonctionnalité de regroupement à implémenter")

    def show_history(self):
        """Affiche l'historique du fichier"""
        QMessageBox.information(self, "Info", "Historique à implémenter")

    def import_manual_file(self):
        """Importe un fichier manuellement"""
        QMessageBox.information(self, "Info", "Import manuel à implémenter")

    def open_transformation_config(self):
        """Ouvre la fenêtre de configuration des transformations"""
        try:
            dialog = TransformationConfigDialog(self)
            dialog.transformation_saved.connect(self.on_transformation_saved)
            dialog.exec()
        except Exception as e:
            logger.error(f"Erreur ouverture config transformations: {e}")
            QMessageBox.critical(self, "Erreur", f"Erreur:\n{str(e)}")

    def on_transformation_saved(self, supplier_id: str):
        """Appelé quand une transformation est sauvegardée"""
        logger.info(f"Transformation sauvegardée pour {supplier_id}")
        self.statusBar.showMessage(f"Transformation sauvegardée pour {supplier_id}", 3000)

    def open_order_analysis(self):
        """Ouvre la fenêtre d'analyse des commandes"""
        try:
            from app.ui.order_analysis_window import OrderAnalysisWindow

            self.analysis_window = OrderAnalysisWindow(self)
            self.analysis_window.show()

        except Exception as e:
            logger.error(f"Erreur ouverture analyse commandes: {e}")
            QMessageBox.critical(self, "Erreur", f"Erreur:\n{str(e)}")

    def open_suppliers_manager(self):
        """Ouvre la fenêtre de gestion des fournisseurs"""
        try:
            from app.ui.suppliers_manager_dialog_v2 import SuppliersManagerDialog

            dialog = SuppliersManagerDialog(self)
            dialog.suppliers_updated.connect(self.on_suppliers_updated)
            dialog.exec()

        except Exception as e:
            logger.error(f"Erreur ouverture gestion fournisseurs: {e}")
            QMessageBox.critical(self, "Erreur", f"Erreur:\n{str(e)}")

    def open_settings(self):
        """Ouvre la fenêtre des réglages"""
        try:
            from app.ui.settings_dialog import SettingsDialog

            dialog = SettingsDialog(self)
            dialog.settings_updated.connect(self.on_settings_updated)
            dialog.exec()

        except Exception as e:
            logger.error(f"Erreur ouverture réglages: {e}")
            QMessageBox.critical(self, "Erreur", f"Erreur:\n{str(e)}")

    def on_settings_updated(self):
        """Appelé quand les réglages sont modifiés"""
        logger.info("Réglages mis à jour")
        self.statusBar.showMessage("Réglages mis à jour", 3000)
        # Reconfigurer le rafraîchissement automatique
        self.setup_auto_refresh()

    def setup_auto_refresh(self):
        """Configure le rafraîchissement automatique selon les préférences"""
        from app.services.user_preferences import get_refresh_interval

        interval = get_refresh_interval()

        # Arrêter le timer existant
        if self.refresh_timer.isActive():
            self.refresh_timer.stop()

        if interval > 0:
            # Démarrer le timer avec l'intervalle en millisecondes
            self.refresh_timer.start(interval * 1000)
            logger.info(f"Rafraîchissement automatique activé: toutes les {interval} secondes")
        else:
            logger.info("Rafraîchissement automatique désactivé")

    def auto_refresh_files(self):
        """Rafraîchit automatiquement la liste des fichiers (appelé par le timer)"""
        logger.info("Rafraîchissement automatique de la liste des fichiers...")
        self.refresh_files_list()

    def on_suppliers_updated(self):
        """Appelé quand les fournisseurs sont modifiés"""
        logger.info("Liste fournisseurs mise à jour")
        self.statusBar.showMessage("Liste fournisseurs mise à jour", 3000)
        # Recharger la grille des fournisseurs
        self.load_suppliers_grid()

    def archive_file(self):
        """Archive les fichiers cochés vers le dossier 'old' sur le serveur FTP"""
        # Récupérer les fichiers cochés
        checked_files = self.get_checked_files()

        if not checked_files:
            QMessageBox.warning(self, "Attention", "Aucun fichier coché à archiver")
            return

        # Confirmation
        file_list = "\n".join([f"• {f['filename']}" for f in checked_files])
        reply = QMessageBox.question(
            self,
            "Confirmation",
            f"Confirmer le déplacement des fichiers dans le dossier des fichiers déjà traités ?\n\n"
            f"Fichiers à archiver ({len(checked_files)}):\n{file_list}\n\n"
            f"Ces fichiers seront déplacés vers le dossier 'old' sur le serveur FTP.",
            QMessageBox.Yes | QMessageBox.No
        )

        if reply != QMessageBox.Yes:
            return

        try:
            self.statusBar.showMessage(f"Archivage de {len(checked_files)} fichier(s) en cours...")

            # Récupérer la config FTP depuis .env
            host = os.getenv("FTP_HOST")
            port = int(os.getenv("FTP_PORT", 22))
            username = os.getenv("FTP_USERNAME")
            password = os.getenv("FTP_PASSWORD")

            # Se connecter au serveur FTP
            fetcher = FTPFetcher(host, port, username, password, use_sftp=True)

            if not fetcher.connect():
                QMessageBox.critical(self, "Erreur", "Impossible de se connecter au serveur FTP")
                return

            # Archiver chaque fichier coché
            success_count = 0
            failed_files = []

            for file_info in checked_files:
                filename = file_info['filename']
                full_path = file_info['full_path']

                logger.info(f"Tentative d'archivage de: {filename} ({full_path})")

                try:
                    # Déplacer vers old/
                    if fetcher.move_to_archive(full_path, "old"):
                        success_count += 1
                        logger.info(f"✓ Fichier archivé sur FTP: {filename}")
                    else:
                        failed_files.append(filename)
                        logger.error(f"✗ Échec d'archivage pour: {filename} (move_to_archive a retourné False)")
                except Exception as e:
                    failed_files.append(filename)
                    logger.error(f"✗ Exception lors de l'archivage de {filename}: {e}")
                    import traceback
                    traceback.print_exc()

            fetcher.disconnect()

            # Afficher le résultat
            if success_count == len(checked_files):
                self.statusBar.showMessage(f"✅ {success_count} fichier(s) archivé(s) avec succès", 5000)
                QMessageBox.information(
                    self,
                    "Succès",
                    f"Tous les fichiers ont été archivés avec succès!\n\n"
                    f"{success_count} fichier(s) déplacé(s) vers le dossier 'old'"
                )
            elif success_count > 0:
                failed_list = "\n".join([f"• {f}" for f in failed_files])
                self.statusBar.showMessage(f"⚠️ {success_count}/{len(checked_files)} fichier(s) archivé(s)", 5000)
                QMessageBox.warning(
                    self,
                    "Archivage partiel",
                    f"{success_count} fichier(s) archivé(s) avec succès.\n\n"
                    f"Échecs ({len(failed_files)}):\n{failed_list}"
                )
            else:
                self.statusBar.showMessage("❌ Échec de l'archivage", 5000)
                QMessageBox.critical(
                    self,
                    "Erreur",
                    "Aucun fichier n'a pu être archivé.\n"
                    "Vérifiez les logs pour plus de détails."
                )

            # Réinitialiser le filtre fournisseur et rafraîchir la liste
            self.clear_supplier_filter()
            self.refresh_files_list()

        except Exception as e:
            logger.error(f"Erreur archivage fichiers: {e}")
            QMessageBox.critical(self, "Erreur", f"Erreur:\n{str(e)}")
            self.statusBar.showMessage("❌ Erreur d'archivage")

    def print_selection(self):
        """Génère un PDF et ouvre l'aperçu avant impression Windows"""
        if not self.selected_supplier_filter:
            QMessageBox.warning(self, "Attention", "Veuillez sélectionner un fournisseur pour utiliser la fonction d'impression")
            return

        try:
            # Récupérer les fichiers filtrés affichés dans le tableau
            filtered_files = []
            for row in range(self.files_table.rowCount()):
                filename = self.files_table.item(row, 0).text()
                full_path = self.files_table.item(row, 2).text()  # Colonne 2 = chemin complet
                filtered_files.append({'filename': filename, 'full_path': full_path})

            if not filtered_files:
                QMessageBox.warning(self, "Attention", "Aucun fichier à imprimer")
                return

            logger.info(f"Impression de {len(filtered_files)} fichier(s) pour {self.selected_supplier_filter}")

            # Récupérer la configuration d'impression du fournisseur
            response = supabase_client.client.table('suppliers').select('*').eq('file_filter_slug', self.selected_supplier_filter).execute()
            if not response.data:
                QMessageBox.warning(self, "Erreur", "Impossible de trouver les paramètres du fournisseur")
                return

            supplier_data = response.data[0]
            print_config = supplier_data.get('print_config', {})

            # Extraire les paramètres
            columns_to_remove = print_config.get('columns_to_remove', [])

            # Gérer les préfixes (nouveau format: tableau, ancien format: string unique)
            prefixes_to_remove = print_config.get('prefixes_to_remove', [])
            if not prefixes_to_remove:
                # Rétrocompatibilité avec l'ancien format
                old_prefix = print_config.get('prefix_to_remove', '')
                if old_prefix:
                    prefixes_to_remove = [old_prefix]

            add_date = print_config.get('add_date', False)
            split_files = print_config.get('split_files', False)
            paper_format = print_config.get('paper_format', 'A4')

            logger.info(f"Configuration d'impression: colonnes={columns_to_remove}, préfixes={prefixes_to_remove}, date={add_date}, split={split_files}, format={paper_format}")

            # Connexion FTP pour télécharger les fichiers
            ftp_host = os.getenv("FTP_HOST")
            ftp_port = int(os.getenv("FTP_PORT", 22))
            ftp_user = os.getenv("FTP_USERNAME")
            ftp_pass = os.getenv("FTP_PASSWORD")

            fetcher = FTPFetcher(ftp_host, ftp_port, ftp_user, ftp_pass, use_sftp=True)
            fetcher.connect()

            # Télécharger tous les fichiers CSV et les garder séparés
            all_dataframes = []
            temp_files = []

            import tempfile

            for file_info in filtered_files:
                full_path = file_info['full_path']
                filename = file_info['filename']

                # Créer un fichier temporaire
                tmp_fd, tmp_path = tempfile.mkstemp(suffix='.csv', prefix='print_')
                os.close(tmp_fd)
                temp_files.append(tmp_path)

                # Télécharger le fichier
                success = fetcher.download_file(full_path, tmp_path)
                if not success:
                    logger.error(f"Échec du téléchargement de {filename}")
                    continue

                # Lire le CSV SANS en-tête (tous les fichiers ont le même format, aucun n'a d'en-tête)
                try:
                    df = pd.read_csv(tmp_path, sep=';', encoding='utf-8', dtype=str, header=None)
                except:
                    try:
                        df = pd.read_csv(tmp_path, sep=',', encoding='utf-8', dtype=str, header=None)
                    except:
                        df = pd.read_csv(tmp_path, sep=';', encoding='latin-1', dtype=str, header=None)

                # Mapper les clés de colonnes vers les indices (position dans le CSV sans en-tête)
                column_index_mapping = {
                    'ref': 0, 'qty': 1, 'designation': 2, 'price': 3,
                    'client': 4, 'ean13': 5, 'order': 6
                }

                # Identifier les colonnes à supprimer par leur indice
                columns_to_drop_for_file = []
                for col_key in columns_to_remove:
                    if col_key in column_index_mapping:
                        col_idx = column_index_mapping[col_key]
                        if col_idx < len(df.columns):
                            columns_to_drop_for_file.append(col_idx)

                # Supprimer les colonnes configurées pour ce fichier
                if columns_to_drop_for_file:
                    columns_to_drop_for_file.sort(reverse=True)
                    for col_idx in columns_to_drop_for_file:
                        df = df.drop(columns=[col_idx])

                # Supprimer les préfixes pour ce fichier (tous les préfixes configurés)
                if prefixes_to_remove and len(df.columns) > 0:
                    first_col = df.columns[0]
                    def remove_prefixes(value):
                        """Supprime le premier préfixe trouvé au début de la valeur"""
                        if not value:
                            return value
                        value_str = str(value)
                        for prefix in prefixes_to_remove:
                            if value_str.startswith(prefix):
                                return value_str[len(prefix):]  # Supprimer le préfixe
                        return value_str

                    df[first_col] = df[first_col].apply(remove_prefixes)

                # Trier ce fichier individuellement SEULEMENT si split_files est activé
                # (sinon le tri global sera appliqué après la fusion)
                if split_files and len(df.columns) > 0:
                    first_col = df.columns[0]
                    df = df.sort_values(by=first_col)
                    logger.debug(f"Tri individuel appliqué au fichier {filename}")

                all_dataframes.append(df)
                logger.info(f"Fichier {filename} traité: {len(df)} lignes, {len(df.columns)} colonnes")

            fetcher.disconnect()

            # Fusionner tous les DataFrames
            if not all_dataframes:
                QMessageBox.warning(self, "Erreur", "Aucune donnée à imprimer")
                return

            # Déterminer le mode de fusion selon le paramètre split_files
            if split_files:
                # MODE SÉPARÉ: Garder les fichiers séparés avec lignes vides
                logger.info(f"Mode séparé (split_files=True) : séparation des fichiers avec lignes vides")
                merged_parts = []
                for i, df in enumerate(all_dataframes):
                    merged_parts.append(df)

                    # Ajouter une ligne vide entre les fichiers (sauf après le dernier)
                    if i < len(all_dataframes) - 1:
                        # Créer une ligne vide avec le même nombre de colonnes
                        empty_row = pd.DataFrame([[''] * len(df.columns)], columns=df.columns)
                        merged_parts.append(empty_row)

                merged_df = pd.concat(merged_parts, ignore_index=True)
                logger.info(f"Total après fusion: {len(merged_df)} lignes (avec {len(all_dataframes)} fichier(s) et {len(all_dataframes)-1} ligne(s) vide(s))")
                logger.info(f"✓ Chaque fichier a été traité individuellement (pas de tri global)")
            else:
                # MODE FUSIONNÉ: Fusion complète avec tri global
                logger.info(f"Mode fusionné (split_files=False) : fusion complète avec tri global par référence")
                merged_df = pd.concat(all_dataframes, ignore_index=True)

                # Tri global par la première colonne (référence)
                if len(merged_df.columns) > 0:
                    first_col = merged_df.columns[0]
                    merged_df = merged_df.sort_values(by=first_col)
                    logger.info(f"✓ Tri global appliqué sur la colonne '{first_col}'")

                logger.info(f"Total après fusion: {len(merged_df)} lignes (fusion complète)")

            logger.info(f"Colonnes du DataFrame fusionné: {list(merged_df.columns)}")

            # Générer le PDF avec reportlab
            from reportlab.lib.pagesizes import A4, A3, portrait, landscape
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import mm

            output_path = Path.home() / "Downloads" / f"impression_{self.selected_supplier_filter}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            output_path.parent.mkdir(exist_ok=True)

            # Configurer la page - TOUJOURS A4 PORTRAIT
            pagesize = portrait(A4)

            # Calculer la largeur disponible
            page_width = pagesize[0] - 20*mm  # Largeur page - marges
            page_height = pagesize[1] - 20*mm

            doc = SimpleDocTemplate(
                str(output_path),
                pagesize=pagesize,
                leftMargin=10*mm,
                rightMargin=10*mm,
                topMargin=10*mm,
                bottomMargin=10*mm
            )

            # Préparer les données du tableau
            data = []
            for _, row in merged_df.iterrows():
                data.append([str(val) if pd.notna(val) else '' for val in row])

            # Calculer les largeurs de colonnes en fonction du contenu
            from reportlab.pdfbase.pdfmetrics import stringWidth
            num_cols = len(data[0]) if data else 3

            # Calculer la largeur maximale pour chaque colonne
            col_widths = []
            font_size = 10
            font_name = 'Helvetica'

            for col_idx in range(num_cols):
                max_width = 0
                for row in data:
                    if col_idx < len(row):
                        text = str(row[col_idx])
                        # Calculer la largeur du texte en points
                        text_width = stringWidth(text, font_name, font_size)
                        max_width = max(max_width, text_width)

                # Ajouter 2 caractères de marge (environ 10-12 points)
                col_widths.append(max_width + stringWidth('XX', font_name, font_size))

            # Vérifier que le total ne dépasse pas la largeur de page
            total_width = sum(col_widths)
            if total_width > page_width:
                # Réduire proportionnellement toutes les colonnes
                scale_factor = page_width / total_width
                col_widths = [w * scale_factor for w in col_widths]

            # Créer le tableau avec largeurs de colonnes ajustées
            table = Table(data, repeatRows=0, colWidths=col_widths)

            # Style du tableau
            style = TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), font_size),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                # Alignements par colonne
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),    # Colonne A (références) : gauche
                ('ALIGN', (1, 0), (1, -1), 'CENTER'),  # Colonne B (quantité) : centre
                ('ALIGN', (2, 0), (2, -1), 'LEFT'),    # Colonne C (désignation) : gauche
                ('ALIGN', (3, 0), (3, -1), 'LEFT'),    # Colonne D (nom client) : gauche
                ('ALIGN', (4, 0), (4, -1), 'LEFT'),    # Colonne E (ref commande) : gauche
            ])
            table.setStyle(style)

            # Construire le document
            elements = [table]

            # Ajouter la date si configuré
            if add_date and len(merged_df.columns) >= 3:
                from reportlab.lib.styles import ParagraphStyle
                from reportlab.lib.enums import TA_CENTER

                date_value = "Le " + datetime.now().strftime('%d/%m/%y')
                elements.append(Spacer(1, 5*mm))

                # Style personnalisé pour centrer la date
                centered_style = ParagraphStyle(
                    'Centered',
                    parent=getSampleStyleSheet()['Normal'],
                    alignment=TA_CENTER,
                    fontSize=12,
                    fontName='Helvetica-Bold'
                )

                date_para = Paragraph(date_value, centered_style)
                elements.append(date_para)
                logger.info(f"✓ Date '{date_value}' ajoutée au PDF (centrée)")

            doc.build(elements)

            # Nettoyer les fichiers temporaires
            for tmp_file in temp_files:
                try:
                    os.unlink(tmp_file)
                except:
                    pass

            logger.info(f"PDF créé: {output_path}")

            # Ouvrir avec l'aperçu avant impression Windows
            import platform
            import subprocess

            if platform.system() == 'Windows':
                # Utiliser le verbe 'printto' pour ouvrir l'aperçu avant impression
                os.startfile(str(output_path), 'open')
                logger.info("PDF ouvert pour aperçu avant impression")
            elif platform.system() == 'Darwin':  # macOS
                subprocess.run(['open', str(output_path)])
            else:  # Linux
                subprocess.run(['xdg-open', str(output_path)])

        except Exception as e:
            logger.error(f"Erreur lors de la génération du PDF: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Erreur", f"Erreur lors de la génération du PDF:\n{str(e)}")

    def export_data(self):
        """Exporte les données filtrées vers un fichier Excel ou CSV selon la configuration Import"""
        if not self.selected_supplier_filter:
            QMessageBox.warning(self, "Attention", "Veuillez sélectionner un fournisseur pour exporter les données")
            return

        try:
            # Récupérer uniquement les fichiers cochés
            filtered_files = self.get_checked_files()

            if not filtered_files:
                QMessageBox.warning(self, "Attention", "Aucun fichier coché à exporter")
                return

            logger.info(f"Export de {len(filtered_files)} fichier(s) pour {self.selected_supplier_filter}")

            # Récupérer la configuration Import du fournisseur
            response = supabase_client.client.table('suppliers').select('*').eq('file_filter_slug', self.selected_supplier_filter).execute()
            if not response.data:
                QMessageBox.warning(self, "Erreur", "Impossible de trouver les paramètres du fournisseur")
                return

            supplier_data = response.data[0]
            supplier_name = supplier_data.get('name', self.selected_supplier_filter)
            import_config = supplier_data.get('import_config', {})

            # Extraire les paramètres d'import
            output_format = import_config.get('output_format', 'xlsx')
            has_header = import_config.get('has_header', False)
            leading_zeros = import_config.get('leading_zeros', False)
            add_output_header = import_config.get('add_output_header', False)
            header_type = import_config.get('header_type', 'Texte fixe')
            header_content = import_config.get('header_content', '')
            merge_duplicates = import_config.get('merge_duplicates', False)
            columns_to_remove = import_config.get('columns_to_remove', [])
            prefixes_to_remove = import_config.get('prefixes_to_remove', [])
            output_filename_pattern = import_config.get('output_filename', '{supplier}_{date}')
            date_format = import_config.get('date_format', 'jj-mm-aa (15-01-25)')

            logger.info(f"Configuration d'import: format={output_format}, colonnes={columns_to_remove}, préfixes={prefixes_to_remove}")
            logger.info(f"En-tête: add_output_header={add_output_header}, header_type={header_type}, header_content='{header_content}'")

            # Connexion FTP pour télécharger les fichiers
            ftp_host = os.getenv("FTP_HOST")
            ftp_port = int(os.getenv("FTP_PORT", 22))
            ftp_user = os.getenv("FTP_USERNAME")
            ftp_pass = os.getenv("FTP_PASSWORD")

            fetcher = FTPFetcher(ftp_host, ftp_port, ftp_user, ftp_pass, use_sftp=True)
            fetcher.connect()

            # Télécharger tous les fichiers CSV et les garder séparés
            all_dataframes = []
            temp_files = []

            import tempfile

            for file_info in filtered_files:
                full_path = file_info['full_path']
                filename = file_info['filename']

                # Créer un fichier temporaire
                tmp_fd, tmp_path = tempfile.mkstemp(suffix='.csv', prefix='export_')
                os.close(tmp_fd)
                temp_files.append(tmp_path)

                # Télécharger le fichier
                success = fetcher.download_file(full_path, tmp_path)
                if not success:
                    logger.error(f"Échec du téléchargement de {filename}")
                    continue

                # Lire le CSV avec ou sans en-tête selon la configuration
                try:
                    if has_header:
                        df = pd.read_csv(tmp_path, sep=';', encoding='utf-8', dtype=str)
                    else:
                        df = pd.read_csv(tmp_path, sep=';', encoding='utf-8', dtype=str, header=None)
                    logger.debug(f"CSV lu avec sep=';' - {len(df.columns)} colonnes, {len(df)} lignes")
                except:
                    try:
                        if has_header:
                            df = pd.read_csv(tmp_path, sep=',', encoding='utf-8', dtype=str)
                        else:
                            df = pd.read_csv(tmp_path, sep=',', encoding='utf-8', dtype=str, header=None)
                        logger.debug(f"CSV lu avec sep=',' - {len(df.columns)} colonnes, {len(df)} lignes")
                    except:
                        if has_header:
                            df = pd.read_csv(tmp_path, sep=';', encoding='latin-1', dtype=str)
                        else:
                            df = pd.read_csv(tmp_path, sep=';', encoding='latin-1', dtype=str, header=None)
                        logger.debug(f"CSV lu avec sep=';' latin-1 - {len(df.columns)} colonnes, {len(df)} lignes")

                all_dataframes.append(df)

            # Fermer la connexion FTP
            fetcher.disconnect()

            # Nettoyer les fichiers temporaires
            for tmp_file in temp_files:
                try:
                    os.remove(tmp_file)
                except:
                    pass

            # Fusionner tous les DataFrames
            if not all_dataframes:
                QMessageBox.warning(self, "Attention", "Aucune donnée à exporter")
                return

            merged_df = pd.concat(all_dataframes, ignore_index=True)
            logger.info(f"Fusion terminée: {len(merged_df)} lignes au total")

            # Appliquer les transformations selon la configuration d'import
            # Mapper les clés de colonnes vers les indices (si pas d'en-tête)
            if not has_header:
                column_index_mapping = {
                    'ref': 0, 'qty': 1, 'designation': 2, 'price': 3,
                    'client': 4, 'ean13': 5, 'order': 6
                }

                # Identifier les colonnes à supprimer par leur indice
                columns_to_drop = []
                for col_key in columns_to_remove:
                    if col_key in column_index_mapping:
                        col_idx = column_index_mapping[col_key]
                        if col_idx < len(merged_df.columns):
                            columns_to_drop.append(col_idx)

                # Supprimer les colonnes configurées
                if columns_to_drop:
                    columns_to_drop.sort(reverse=True)
                    for col_idx in columns_to_drop:
                        merged_df = merged_df.drop(columns=[col_idx])

                # Supprimer les préfixes de la première colonne
                if prefixes_to_remove and len(merged_df.columns) > 0:
                    first_col = merged_df.columns[0]
                    def remove_prefixes(value):
                        """Supprime le premier préfixe trouvé au début de la valeur"""
                        if pd.isna(value):
                            return value
                        value_str = str(value)
                        for prefix in prefixes_to_remove:
                            if value_str.startswith(prefix):
                                return value_str[len(prefix):]
                        return value_str

                    merged_df[first_col] = merged_df[first_col].apply(remove_prefixes)

                # Conserver les zéros de tête si configuré
                if leading_zeros and len(merged_df.columns) > 0:
                    first_col = merged_df.columns[0]
                    # Formater comme texte pour conserver les zéros
                    merged_df[first_col] = merged_df[first_col].astype(str)

            else:
                # Avec en-tête, utiliser les noms de colonnes
                for col_key in columns_to_remove:
                    if col_key in merged_df.columns:
                        merged_df = merged_df.drop(columns=[col_key])

                # Supprimer les préfixes de la colonne 'ref' si elle existe
                if prefixes_to_remove and 'ref' in merged_df.columns:
                    def remove_prefixes(value):
                        if pd.isna(value):
                            return value
                        value_str = str(value)
                        for prefix in prefixes_to_remove:
                            if value_str.startswith(prefix):
                                return value_str[len(prefix):]
                        return value_str

                    merged_df['ref'] = merged_df['ref'].apply(remove_prefixes)

                # Conserver les zéros de tête si configuré
                if leading_zeros and 'ref' in merged_df.columns:
                    merged_df['ref'] = merged_df['ref'].astype(str)

            # Fusionner les doublons si configuré
            if merge_duplicates and len(merged_df.columns) >= 2:
                # Grouper par la première colonne (ref) et sommer la deuxième (qty)
                first_col = merged_df.columns[0]
                second_col = merged_df.columns[1]

                # Convertir la colonne qty en numérique pour la somme
                merged_df[second_col] = pd.to_numeric(merged_df[second_col], errors='coerce').fillna(0).astype(int)

                # Grouper et sommer
                merged_df = merged_df.groupby(first_col, as_index=False).agg({
                    col: 'first' if col != second_col else 'sum'
                    for col in merged_df.columns if col != first_col
                })

            # Trier par la première colonne (références) en ordre alphabétique
            if len(merged_df.columns) > 0:
                first_col = merged_df.columns[0]
                merged_df = merged_df.sort_values(by=first_col, ascending=True)
                merged_df = merged_df.reset_index(drop=True)
                logger.info(f"Données triées par la colonne '{first_col}'")

            # Ajouter un en-tête si configuré
            if add_output_header:
                if (header_type == 'Texte fixe' or header_type == 'Texte + date du jour') and header_content:
                    # Remplacer le placeholder {date} par la date formatée
                    from datetime import datetime
                    now = datetime.now()

                    # Formater la date selon le format choisi
                    if 'jj-mm-aa' in date_format:
                        date_str_header = now.strftime("%d-%m-%y")
                    elif 'jj-mm-aaaa' in date_format:
                        date_str_header = now.strftime("%d-%m-%Y")
                    elif 'aaaa-mm-jj' in date_format:
                        date_str_header = now.strftime("%Y-%m-%d")
                    elif 'jjmmaa' in date_format:
                        date_str_header = now.strftime("%d%m%y")
                    elif 'jjmmaaaa' in date_format:
                        date_str_header = now.strftime("%d%m%Y")
                    else:
                        date_str_header = now.strftime("%d-%m-%y")

                    # Remplacer {date} dans le contenu de l'en-tête
                    processed_header_content = header_content.replace('{date}', date_str_header)

                    # Détecter si le texte contient des séparateurs ';'
                    if ';' in processed_header_content:
                        # Séparer les en-têtes par le séparateur ';'
                        header_values = processed_header_content.split(';')
                        # Compléter avec des chaînes vides si nécessaire
                        while len(header_values) < len(merged_df.columns):
                            header_values.append('')
                        # Tronquer si trop d'en-têtes
                        header_row = header_values[:len(merged_df.columns)]
                    else:
                        # Texte fixe sans séparateur : mettre uniquement dans la première colonne
                        header_row = [processed_header_content] + [''] * (len(merged_df.columns) - 1)

                    header_df = pd.DataFrame([header_row], columns=merged_df.columns)
                    merged_df = pd.concat([header_df, merged_df], ignore_index=True)
                    logger.info(f"En-tête ajouté: {header_row}")
                elif header_type == 'Nom de colonnes':
                    # Utiliser les noms de colonnes comme en-tête
                    # Si pas d'en-tête, créer des noms de colonnes par défaut
                    if not has_header:
                        column_names = ['Ref', 'Qty', 'Designation', 'Price', 'Client', 'EAN13', 'Order']
                        # Ajuster au nombre de colonnes réelles
                        column_names = column_names[:len(merged_df.columns)]
                        merged_df.columns = column_names

            # Générer le nom de fichier selon le pattern
            from datetime import datetime
            now = datetime.now()

            # Formater la date selon le format choisi
            if 'jj-mm-aa' in date_format:
                date_str = now.strftime("%d-%m-%y")
            elif 'jj-mm-aaaa' in date_format:
                date_str = now.strftime("%d-%m-%Y")
            elif 'aaaa-mm-jj' in date_format:
                date_str = now.strftime("%Y-%m-%d")
            elif 'jjmmaa' in date_format:
                date_str = now.strftime("%d%m%y")
            elif 'jjmmaaaa' in date_format:
                date_str = now.strftime("%d%m%Y")
            else:
                date_str = now.strftime("%d-%m-%y")

            # Remplacer les placeholders dans le nom de fichier
            filename = output_filename_pattern.replace('{supplier}', supplier_name).replace('{date}', date_str)

            # Ajouter l'extension selon le format
            if output_format == 'csv':
                filename += '.csv'
            else:
                filename += '.xlsx'

            # Obtenir le dossier de sortie configuré
            from app.services.user_preferences import get_output_folder
            output_folder = get_output_folder()
            output_path = output_folder / filename

            # Sauvegarder le fichier selon le format
            if output_format == 'csv':
                merged_df.to_csv(output_path, index=False, header=add_output_header and header_type == 'Nom de colonnes', sep=';', encoding='utf-8')
                logger.success(f"Fichier CSV exporté: {output_path}")
            else:
                # Créer le fichier Excel avec openpyxl pour pouvoir ajuster les largeurs de colonnes
                from openpyxl import Workbook
                from openpyxl.utils.dataframe import dataframe_to_rows

                wb = Workbook()
                ws = wb.active

                # Écrire les données
                for r_idx, row in enumerate(dataframe_to_rows(merged_df, index=False, header=False), 1):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx, column=c_idx, value=value)

                # Ajuster la largeur des colonnes en fonction du contenu
                for col_idx, column in enumerate(ws.columns, 1):
                    max_length = 0
                    column_letter = ws.cell(row=1, column=col_idx).column_letter

                    for cell in column:
                        try:
                            if cell.value:
                                # Calculer la longueur du contenu
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass

                    # Ajouter 2 caractères de marge
                    adjusted_width = max_length + 2
                    ws.column_dimensions[column_letter].width = adjusted_width

                # Sauvegarder le fichier
                wb.save(output_path)
                logger.success(f"Fichier Excel exporté avec colonnes auto-ajustées: {output_path}")

            # Libérer la mémoire du DataFrame pour éviter les conflits OLE
            del merged_df

            # Ouvrir le fichier avec Excel (nouvelle instance isolée)
            import subprocess

            excel_paths = [
                r'C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE',
                r'C:\Program Files (x86)\Microsoft Office\root\Office16\EXCEL.EXE',
                r'C:\Program Files\Microsoft Office\Office16\EXCEL.EXE',
                r'C:\Program Files (x86)\Microsoft Office\Office16\EXCEL.EXE'
            ]

            excel_opened = False
            for excel_path in excel_paths:
                if Path(excel_path).exists():
                    try:
                        # Utiliser /x pour ouvrir une nouvelle instance d'Excel isolée
                        # DETACHED_PROCESS évite les conflits OLE
                        process = subprocess.Popen(
                            [excel_path, '/x', str(output_path)],
                            creationflags=subprocess.CREATE_NEW_PROCESS_GROUP | subprocess.DETACHED_PROCESS,
                            stdout=subprocess.DEVNULL,
                            stderr=subprocess.DEVNULL,
                            close_fds=True
                        )
                        logger.info(f"Fichier ouvert avec Excel dans une nouvelle instance: {output_path}")
                        excel_opened = True
                        break
                    except Exception as e:
                        logger.error(f"Erreur lors de l'ouverture avec Excel ({excel_path}): {e}")
                        continue

            if not excel_opened:
                # Fallback: ouvrir avec l'application par défaut
                try:
                    os.startfile(str(output_path))
                    logger.info(f"Fichier ouvert avec l'application par défaut: {output_path}")
                except Exception as e:
                    logger.error(f"Erreur lors de l'ouverture du fichier: {e}")
                    QMessageBox.information(
                        self,
                        "Export réussi",
                        f"Fichier exporté:\n{output_path}\n\nVeuillez l'ouvrir manuellement."
                    )
                    return

            logger.success("Export terminé avec succès")

        except Exception as e:
            logger.error(f"Erreur lors de l'export: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(
                self,
                "Erreur",
                f"Erreur lors de l'export:\n{str(e)}"
            )

    def open_file(self):
        """Ouvre le fichier fusionné dans Excel (comme pour l'impression mais sans PDF)"""
        if not self.selected_supplier_filter:
            QMessageBox.warning(self, "Attention", "Veuillez sélectionner un fournisseur pour ouvrir le fichier")
            return

        try:
            # Récupérer uniquement les fichiers cochés
            filtered_files = self.get_checked_files()

            if not filtered_files:
                QMessageBox.warning(self, "Attention", "Aucun fichier coché à ouvrir")
                return

            logger.info(f"Ouverture de {len(filtered_files)} fichier(s) pour {self.selected_supplier_filter}")

            # Récupérer la configuration d'affichage du fournisseur
            response = supabase_client.client.table('suppliers').select('*').eq('file_filter_slug', self.selected_supplier_filter).execute()
            if not response.data:
                QMessageBox.warning(self, "Erreur", "Impossible de trouver les paramètres du fournisseur")
                return

            supplier_data = response.data[0]
            display_config = supplier_data.get('display_config', {})

            # Extraire les paramètres d'affichage
            columns_to_remove = display_config.get('columns_to_remove', [])
            prefixes_to_remove = display_config.get('prefixes_to_remove', [])
            add_date = display_config.get('add_date', False)
            split_files = display_config.get('split_files', False)

            logger.info(f"Configuration d'affichage: colonnes={columns_to_remove}, préfixes={prefixes_to_remove}, date={add_date}, split={split_files}")

            # Connexion FTP pour télécharger les fichiers
            ftp_host = os.getenv("FTP_HOST")
            ftp_port = int(os.getenv("FTP_PORT", 22))
            ftp_user = os.getenv("FTP_USERNAME")
            ftp_pass = os.getenv("FTP_PASSWORD")

            fetcher = FTPFetcher(ftp_host, ftp_port, ftp_user, ftp_pass, use_sftp=True)
            fetcher.connect()

            # Télécharger tous les fichiers CSV et les garder séparés
            all_dataframes = []
            temp_files = []

            import tempfile

            for file_info in filtered_files:
                full_path = file_info['full_path']
                filename = file_info['filename']

                # Créer un fichier temporaire
                tmp_fd, tmp_path = tempfile.mkstemp(suffix='.csv', prefix='open_')
                os.close(tmp_fd)
                temp_files.append(tmp_path)

                # Télécharger le fichier
                success = fetcher.download_file(full_path, tmp_path)
                if not success:
                    logger.error(f"Échec du téléchargement de {filename}")
                    continue

                # Lire le CSV SANS en-tête
                try:
                    df = pd.read_csv(tmp_path, sep=';', encoding='utf-8', dtype=str, header=None)
                    logger.debug(f"CSV lu avec sep=';' - {len(df.columns)} colonnes, {len(df)} lignes")
                except:
                    try:
                        df = pd.read_csv(tmp_path, sep=',', encoding='utf-8', dtype=str, header=None)
                        logger.debug(f"CSV lu avec sep=',' - {len(df.columns)} colonnes, {len(df)} lignes")
                    except:
                        df = pd.read_csv(tmp_path, sep=';', encoding='latin-1', dtype=str, header=None)
                        logger.debug(f"CSV lu avec sep=';' latin-1 - {len(df.columns)} colonnes, {len(df)} lignes")

                # Mapper les clés de colonnes vers les indices
                column_index_mapping = {
                    'ref': 0, 'qty': 1, 'designation': 2, 'price': 3,
                    'client': 4, 'ean13': 5, 'order': 6
                }

                # Identifier les colonnes à supprimer par leur indice
                columns_to_drop_for_file = []
                for col_key in columns_to_remove:
                    if col_key in column_index_mapping:
                        col_idx = column_index_mapping[col_key]
                        if col_idx < len(df.columns):
                            columns_to_drop_for_file.append(col_idx)

                # Supprimer les colonnes configurées
                if columns_to_drop_for_file:
                    columns_to_drop_for_file.sort(reverse=True)
                    for col_idx in columns_to_drop_for_file:
                        df = df.drop(columns=[col_idx])

                # Supprimer les préfixes
                if prefixes_to_remove and len(df.columns) > 0:
                    first_col = df.columns[0]
                    def remove_prefixes(value):
                        """Supprime le premier préfixe trouvé au début de la valeur"""
                        if not value:
                            return value
                        value_str = str(value)
                        for prefix in prefixes_to_remove:
                            if value_str.startswith(prefix):
                                return value_str[len(prefix):]
                        return value_str

                    df[first_col] = df[first_col].apply(remove_prefixes)

                # Trier ce fichier individuellement SEULEMENT si split_files est activé
                if split_files and len(df.columns) > 0:
                    first_col = df.columns[0]
                    df = df.sort_values(by=first_col)
                    logger.debug(f"Tri individuel appliqué au fichier {filename}")

                all_dataframes.append(df)
                logger.info(f"Fichier {filename} traité: {len(df)} lignes, {len(df.columns)} colonnes")

            fetcher.disconnect()

            # Fusionner tous les DataFrames
            if not all_dataframes:
                QMessageBox.warning(self, "Erreur", "Aucune donnée à ouvrir")
                return

            # Déterminer le mode de fusion selon le paramètre split_files
            if split_files:
                # MODE SÉPARÉ: Garder les fichiers séparés avec lignes vides
                logger.info(f"Mode séparé (split_files=True) : séparation des fichiers avec lignes vides")
                merged_parts = []
                for i, df in enumerate(all_dataframes):
                    merged_parts.append(df)

                    # Ajouter une ligne vide entre les fichiers (sauf après le dernier)
                    if i < len(all_dataframes) - 1:
                        # Créer une ligne vide avec le même nombre de colonnes
                        empty_row = pd.DataFrame([[''] * len(df.columns)], columns=df.columns)
                        merged_parts.append(empty_row)

                merged_df = pd.concat(merged_parts, ignore_index=True)
                logger.info(f"Total après fusion: {len(merged_df)} lignes (avec {len(all_dataframes)} fichier(s) et {len(all_dataframes)-1} ligne(s) vide(s))")
            else:
                # MODE FUSIONNÉ: Fusion complète avec tri global
                logger.info(f"Mode fusionné (split_files=False) : fusion complète avec tri global par référence")
                merged_df = pd.concat(all_dataframes, ignore_index=True)

                # Tri global par la première colonne (référence)
                if len(merged_df.columns) > 0:
                    first_col = merged_df.columns[0]
                    merged_df = merged_df.sort_values(by=first_col)
                    logger.info(f"✓ Tri global appliqué sur la colonne '{first_col}'")

                logger.info(f"Total après fusion: {len(merged_df)} lignes (fusion complète)")

            # Mémoriser le nombre de lignes de données (avant d'ajouter la date)
            data_rows_count = len(merged_df)

            # Ajouter la date si configuré (avec une ligne vide avant)
            if add_date:
                date_value = "Le " + datetime.now().strftime('%d/%m/%y')

                # Ajouter d'abord une ligne vide
                empty_row = [''] * len(merged_df.columns)
                empty_df = pd.DataFrame([empty_row], columns=merged_df.columns)
                merged_df = pd.concat([merged_df, empty_df], ignore_index=True)

                # Puis ajouter la ligne avec la date dans la 3ème colonne (index 2)
                date_row = [''] * len(merged_df.columns)
                if len(date_row) >= 3:
                    date_row[2] = date_value
                    date_df = pd.DataFrame([date_row], columns=merged_df.columns)
                    merged_df = pd.concat([merged_df, date_df], ignore_index=True)
                    logger.info(f"✓ Ligne vide + Date '{date_value}' ajoutées à la fin du fichier (colonne C)")

            # Générer le fichier Excel dans le dossier de sortie configuré
            from app.services.user_preferences import get_output_folder

            date_str = datetime.now().strftime('%d-%m-%y')  # Format jj-mm-aa
            output_folder = get_output_folder()
            output_path = output_folder / f"{self.selected_supplier_filter}-{date_str}-visu.xlsx"
            output_path.parent.mkdir(exist_ok=True, parents=True)

            logger.info(f"DataFrame final: {len(merged_df)} lignes, {len(merged_df.columns)} colonnes")
            logger.debug(f"Colonnes du DataFrame: {list(merged_df.columns)}")

            # Créer le workbook avec openpyxl directement pour un meilleur contrôle de la fermeture
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = 'Sheet1'

            # Écrire les données ligne par ligne et appliquer les bordures
            from openpyxl.styles import Border, Side, Alignment

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            center_alignment = Alignment(horizontal='center', vertical='center')

            for row_idx, row_data in enumerate(merged_df.values, start=1):
                for col_idx, value in enumerate(row_data, start=1):
                    # Convertir les valeurs numériques des colonnes B (qté) et E (code) en nombres
                    if col_idx in [2, 5] and value and str(value).strip():
                        try:
                            # Essayer de convertir en nombre
                            numeric_value = float(value) if '.' in str(value) else int(value)
                            cell = worksheet.cell(row=row_idx, column=col_idx, value=numeric_value)
                        except (ValueError, TypeError):
                            # Si la conversion échoue, garder comme texte
                            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                    else:
                        cell = worksheet.cell(row=row_idx, column=col_idx, value=value)

                    # Appliquer les bordures uniquement sur les lignes de données (pas sur la date)
                    if row_idx <= data_rows_count and value and str(value).strip():
                        cell.border = thin_border

                    # Centrer la colonne B (quantités)
                    if col_idx == 2 and value and str(value).strip():
                        cell.alignment = center_alignment

            # Auto-ajuster la largeur de toutes les colonnes
            for col_idx in range(1, len(merged_df.columns) + 1):
                column_letter = get_column_letter(col_idx)

                # Calculer la largeur max pour cette colonne (uniquement sur les lignes de données)
                max_length = 0
                for row_idx in range(1, data_rows_count + 1):
                    cell = worksheet[f'{column_letter}{row_idx}']
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

                # Pour la colonne C (désignation), ajuster exactement à la largeur du texte
                if col_idx == 3:
                    # Adapter exactement à la longueur du texte le plus long (+ 1 pour une petite marge)
                    adjusted_width = max_length + 1
                else:
                    # Pour les autres colonnes, ajouter une petite marge et limiter à 50
                    adjusted_width = min(max_length + 2, 50)

                worksheet.column_dimensions[column_letter].width = adjusted_width

            logger.info(f"Auto-ajustement de {len(merged_df.columns)} colonnes Excel effectué")

            # Sauvegarder le workbook
            workbook.save(output_path)

            # IMPORTANT: Fermeture complète et nettoyage AVANT d'ouvrir Excel
            workbook.close()
            del workbook
            del worksheet

            # Forcer le garbage collector
            import gc
            gc.collect()

            # Nettoyer les fichiers temporaires
            for tmp_file in temp_files:
                try:
                    os.unlink(tmp_file)
                except:
                    pass

            logger.info(f"Fichier Excel créé: {output_path}")

            # ATTENDRE que tous les processus Python libèrent le fichier
            import time
            time.sleep(0.5)

            # Vérifier que le fichier existe bien
            if not output_path.exists():
                raise FileNotFoundError(f"Le fichier {output_path} n'a pas été créé correctement")

            logger.info(f"Ouverture du fichier Excel...")

            # Ouvrir Excel de manière complètement détachée pour éviter les conflits OLE
            import subprocess
            import platform

            if platform.system() == 'Windows':
                # Essayer d'abord d'ouvrir avec Excel directement avec l'option /x (nouvelle instance)
                excel_paths = [
                    r'C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE',
                    r'C:\Program Files (x86)\Microsoft Office\root\Office16\EXCEL.EXE',
                    r'C:\Program Files\Microsoft Office\Office16\EXCEL.EXE',
                    r'C:\Program Files (x86)\Microsoft Office\Office16\EXCEL.EXE',
                    r'C:\Program Files\Microsoft Office\root\Office15\EXCEL.EXE',
                    r'C:\Program Files (x86)\Microsoft Office\root\Office15\EXCEL.EXE',
                ]

                excel_opened = False
                for excel_path in excel_paths:
                    if Path(excel_path).exists():
                        try:
                            # Méthode 1: Utiliser /x pour forcer une nouvelle instance isolée
                            logger.debug(f"Tentative d'ouverture avec /x: {excel_path}")
                            process = subprocess.Popen(
                                [excel_path, '/x', str(output_path)],
                                creationflags=subprocess.CREATE_NEW_PROCESS_GROUP | subprocess.DETACHED_PROCESS,
                                stdout=subprocess.DEVNULL,
                                stderr=subprocess.DEVNULL,
                                close_fds=True
                            )
                            # Attendre un peu pour voir si ça a marché
                            time.sleep(0.3)
                            if process.poll() is None or process.returncode == 0:
                                logger.info(f"✓ Fichier Excel ouvert avec nouvelle instance: {output_path.name}")
                                excel_opened = True
                                break
                        except Exception as e:
                            logger.debug(f"Échec méthode /x avec {excel_path}: {e}")

                        # Méthode 2: Essayer sans /x si /x a échoué
                        try:
                            logger.debug(f"Tentative d'ouverture sans /x: {excel_path}")
                            subprocess.Popen(
                                [excel_path, str(output_path)],
                                creationflags=subprocess.CREATE_NEW_PROCESS_GROUP | subprocess.DETACHED_PROCESS,
                                stdout=subprocess.DEVNULL,
                                stderr=subprocess.DEVNULL,
                                close_fds=True
                            )
                            logger.info(f"✓ Fichier Excel ouvert: {output_path.name}")
                            excel_opened = True
                            break
                        except Exception as e:
                            logger.debug(f"Échec sans /x avec {excel_path}: {e}")
                            continue

                # Si Excel.exe n'a pas été trouvé, utiliser os.startfile avec gestion du message OLE
                if not excel_opened:
                    logger.info("Excel.exe non trouvé, utilisation de os.startfile")
                    # Simplement ouvrir le fichier et ignorer le message OLE (l'utilisateur cliquera sur OK)
                    os.startfile(str(output_path))
                    logger.info(f"Fichier Excel ouvert (via startfile): {output_path.name}")

            logger.info(f"Fichier préparé: {output_path.name} - {len(filtered_files)} fichiers fusionnés, {len(merged_df)} lignes totales")

        except Exception as e:
            logger.error(f"Erreur lors de l'ouverture du fichier: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Erreur", f"Erreur lors de l'ouverture du fichier:\n{str(e)}")

    def open_supplier_website(self):
        """Ouvre le site web du fournisseur (à paramétrer)"""
        if not self.selected_supplier_filter:
            QMessageBox.warning(self, "Attention", "Veuillez sélectionner un fournisseur")
            return

        # Récupérer les informations du fournisseur
        try:
            response = supabase_client.client.table('suppliers').select('*').eq('file_filter_slug', self.selected_supplier_filter).execute()
            if not response.data:
                QMessageBox.warning(self, "Erreur", "Impossible de trouver le fournisseur")
                return

            supplier_data = response.data[0]

            # Récupérer web_config et website
            web_config = supplier_data.get('web_config', {})
            if isinstance(web_config, dict):
                url_from_config = web_config.get('url', '')
                login_value = web_config.get('login_value', '')
                password_value = web_config.get('password_value', '')
            else:
                url_from_config = ''
                login_value = ''
                password_value = ''

            website_url = supplier_data.get('website', '')

            # Déterminer l'URL à utiliser
            url_to_open = url_from_config or website_url

            if not url_to_open:
                QMessageBox.information(
                    self,
                    "Pas de site web",
                    f"Aucun site web configuré pour le fournisseur {self.selected_supplier_filter}"
                )
                return

            # Décider : automatisation ou simple ouverture ?
            # Si password est renseigné → automatisation (peu importe les autres champs)
            # Sinon → simple ouverture dans le navigateur par défaut
            needs_automation = bool(password_value)

            if needs_automation:
                logger.info("Mode automatisation : mot de passe détecté")

                # Récupérer la configuration du navigateur depuis les préférences
                from app.services.user_preferences import get_browser_config
                browser_config = get_browser_config()
                browser_choice = browser_config.get('browser', 'Navigateur par défaut du système')

                # Mapper le choix vers le type Selenium
                browser_map = {
                    'Google Chrome': 'chrome',
                    'Mozilla Firefox': 'firefox',
                    'Microsoft Edge': 'edge',
                    'Navigateur par défaut du système': 'chrome',
                }
                browser_type = browser_map.get(browser_choice, 'chrome')

                # Lancer l'automatisation dans un thread séparé
                from PySide6.QtCore import QThread, Signal

                class WebAutomationThread(QThread):
                    finished = Signal(bool, str)

                    def __init__(self, web_config, browser_type):
                        super().__init__()
                        self.web_config = web_config
                        self.browser_type = browser_type

                    def run(self):
                        try:
                            from app.services.web_automation import auto_login_supplier
                            success = auto_login_supplier(self.web_config, self.browser_type)
                            if success:
                                self.finished.emit(True, "Connexion automatique réussie!")
                            else:
                                self.finished.emit(False, "Erreur lors de la connexion automatique")
                        except Exception as e:
                            self.finished.emit(False, f"Erreur: {str(e)}")

                def on_automation_finished(success, message):
                    if not success:
                        QMessageBox.warning(self, "Automatisation Web", message)

                self.web_thread = WebAutomationThread(web_config, browser_type)
                self.web_thread.finished.connect(on_automation_finished)
                self.web_thread.start()

                logger.info("Thread d'automatisation web lancé")

            else:
                logger.info("Mode simple : ouverture dans le navigateur par défaut (pas de mot de passe)")
                from app.services.user_preferences import open_url
                success = open_url(url_to_open)
                if success:
                    logger.info(f"Site web ouvert dans le navigateur: {url_to_open}")
                else:
                    QMessageBox.warning(self, "Erreur", "Impossible d'ouvrir le site web")

        except Exception as e:
            logger.error(f"Erreur lors de l'ouverture du site web: {e}")
            QMessageBox.critical(self, "Erreur", f"Erreur:\n{str(e)}")

    def logout(self):
        """Déconnexion"""
        supabase_client.sign_out()
        QMessageBox.information(self, "Déconnexion", "Vous avez été déconnecté")
        sys.exit(0)
